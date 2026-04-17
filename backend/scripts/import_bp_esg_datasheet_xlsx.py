from __future__ import annotations

import argparse
import asyncio
import json
import os
import re
import sys
from dataclasses import asdict, dataclass, field
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import and_, select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

BACKEND_DIR = Path(__file__).resolve().parents[1]
REPO_DIR = BACKEND_DIR.parent
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from app.core.schema_runtime import stamp_database_async
from app.core.security import hash_password
from app.db.models import Base
from app.db.models.boundary import BoundaryDefinition, BoundaryMembership
from app.db.models.company_entity import CompanyEntity
from app.db.models.completeness import (
    DisclosureRequirementStatus,
    RequirementItemDataPoint,
    RequirementItemStatus,
)
from app.db.models.data_point import DataPoint
from app.db.models.evidence import DataPointEvidence, Evidence, EvidenceFile
from app.db.models.mapping import RequirementItemSharedElement
from app.db.models.organization import Organization
from app.db.models.project import MetricAssignment, ReportingProject, ReportingProjectStandard
from app.db.models.requirement_item import RequirementItem
from app.db.models.role_binding import RoleBinding
from app.db.models.shared_element import SharedElement
from app.db.models.standard import DisclosureRequirement, Standard, StandardSection
from app.db.models.unit_reference import UnitReference
from app.db.models.user import User


YEARS = (2019, 2020, 2021, 2022, 2023, 2024)
DATA_SHEETS = (
    "3 Net zero",
    "4 GHG & energy",
    "5 Safety",
    "6 Environment",
    "7 Social",
    "8 Governance",
)
DEFAULT_DATABASE_URL = os.getenv(
    "DATABASE_URL",
    "postgresql+asyncpg://postgres:postgres@127.0.0.1:5432/esgvist",
)
DEFAULT_PASSWORD = os.getenv("BP_PROXY_IMPORT_PASSWORD", "Test1234")
DEFAULT_ORG_NAME = "Atlas Energy Group"
DEFAULT_ORG_LEGAL_NAME = "Atlas Energy Group plc"
DEFAULT_STANDARD_CODE = "ATLAS-ESG-DATASHEET"
DEFAULT_STANDARD_NAME = "Atlas ESG Datasheet (Proxy Import)"
DEFAULT_REPORT_DIR = REPO_DIR / "artifacts" / "imports" / "bp_proxy"
SLUG_RE = re.compile(r"[^a-z0-9]+")
MISSING_MARKERS = {"", "-", "—", "–", "n/a", "na"}
DEFAULT_USER_EMAILS = {
    "admin": "admin@atlasenergy.example.com",
    "manager": "manager@atlasenergy.example.com",
    "reviewer": "reviewer@atlasenergy.example.com",
    "collector": "collector.import@atlasenergy.example.com",
    "auditor": "auditor@atlasenergy.example.com",
}

SECTION_META = {
    "Net zero": {"code": "NET_ZERO", "domain": "net_zero"},
    "GHG & energy": {"code": "GHG_ENERGY", "domain": "emissions_energy"},
    "Safety": {"code": "SAFETY", "domain": "safety"},
    "Environment": {"code": "ENVIRONMENT", "domain": "environment"},
    "Social": {"code": "SOCIAL", "domain": "social"},
    "Governance": {"code": "GOVERNANCE", "domain": "governance"},
}


@dataclass
class ImportConflict:
    code: str
    severity: str
    message: str
    sheet: str | None = None
    row: int | None = None
    context: dict[str, Any] = field(default_factory=dict)


@dataclass
class CellValue:
    raw: str
    numeric_value: float | None
    text_value: str | None
    value_kind: str


@dataclass
class ParsedRow:
    row_number: int
    raw_label: str
    clean_label: str
    footnote_tokens: tuple[str, ...]
    unit: str | None
    depth: int
    is_header: bool
    raw_values: tuple[Any, ...]


@dataclass
class ParsedMetric:
    sheet_name: str
    section_title: str
    row_number: int
    major_group: str
    subgroup: str | None
    path_components: tuple[str, ...]
    tail_components: tuple[str, ...]
    footnote_tokens: tuple[str, ...]
    unit: str | None
    value_mode: str
    values_by_year: dict[int, CellValue]
    shared_element_code: str
    shared_element_name: str
    requirement_item_name: str
    description: str
    disclosure_key: tuple[str, str]

    @property
    def path(self) -> str:
        return " / ".join(self.path_components)


@dataclass
class ParseResult:
    metrics: list[ParsedMetric]
    conflicts: list[ImportConflict]
    section_disclosures: dict[tuple[str, str], str]


@dataclass
class ImportStats:
    created_users: int = 0
    updated_users: int = 0
    created_role_bindings: int = 0
    updated_role_bindings: int = 0
    created_organization: int = 0
    updated_organization: int = 0
    created_entities: int = 0
    updated_entities: int = 0
    created_boundaries: int = 0
    updated_boundaries: int = 0
    created_boundary_memberships: int = 0
    updated_boundary_memberships: int = 0
    created_standard: int = 0
    updated_standard: int = 0
    created_sections: int = 0
    updated_sections: int = 0
    created_disclosures: int = 0
    updated_disclosures: int = 0
    created_units: int = 0
    updated_units: int = 0
    created_shared_elements: int = 0
    updated_shared_elements: int = 0
    created_requirement_items: int = 0
    updated_requirement_items: int = 0
    created_mappings: int = 0
    created_projects: int = 0
    updated_projects: int = 0
    created_project_standards: int = 0
    created_assignments: int = 0
    updated_assignments: int = 0
    created_data_points: int = 0
    updated_data_points: int = 0
    created_item_bindings: int = 0
    created_item_statuses: int = 0
    updated_item_statuses: int = 0
    created_disclosure_statuses: int = 0
    updated_disclosure_statuses: int = 0
    created_evidences: int = 0
    updated_evidences: int = 0
    created_evidence_links: int = 0


@dataclass
class ImportSummary:
    mode: str
    database_url: str
    source_path: str
    report_dir: str
    organization_name: str
    project_ids_by_year: dict[int, int]
    project_names_by_year: dict[int, str]
    counts: dict[str, Any]
    conflicts: list[dict[str, Any]]
    credentials: dict[str, str]
    stats: dict[str, int]


def slugify(value: str) -> str:
    return SLUG_RE.sub("_", value.lower()).strip("_")


def clean_label(label: str, footnote_tokens: tuple[str, ...]) -> str:
    cleaned = re.sub(r"\s+", " ", label.replace("\xa0", " ").strip())
    cleaned = cleaned.rstrip("*").strip()
    if footnote_tokens:
        pattern = r"(?:\s+(?:" + "|".join(re.escape(token) for token in footnote_tokens) + r"))+$"
        cleaned = re.sub(pattern, "", cleaned).strip()
    return cleaned.rstrip("*").strip()


def normalize_unit(value: Any) -> str | None:
    if value is None:
        return None
    cleaned = re.sub(r"\s+", " ", str(value).replace("\xa0", " ").strip())
    return cleaned or None


def parse_footnote_tokens(value: Any) -> tuple[str, ...]:
    if value is None:
        return ()
    tokens = tuple(token for token in re.split(r"\s+", str(value).strip()) if token)
    return tokens


def classify_raw_value(value: Any) -> CellValue | None:
    if value is None:
        return None
    if isinstance(value, bool):
        raw = "true" if value else "false"
        return CellValue(raw=raw, numeric_value=None, text_value=raw, value_kind="text")
    if isinstance(value, (int, float)):
        return CellValue(raw=str(value), numeric_value=float(value), text_value=None, value_kind="number")

    raw = re.sub(r"\s+", " ", str(value).replace("\xa0", " ").strip())
    if raw.lower() in MISSING_MARKERS:
        return None
    try:
        numeric = float(raw)
    except ValueError:
        return CellValue(raw=raw, numeric_value=None, text_value=raw, value_kind="text")
    return CellValue(raw=raw, numeric_value=numeric, text_value=None, value_kind="number")


def build_description(metric: ParsedMetric) -> str:
    details = [
        f"Imported from source sheet '{metric.sheet_name}' row {metric.row_number}.",
        f"Normalized path: {metric.path}.",
    ]
    if metric.unit:
        details.append(f"Unit: {metric.unit}.")
    if metric.footnote_tokens:
        details.append(f"Footnotes: {' '.join(metric.footnote_tokens)}.")
    return " ".join(details)


def load_sheet_rows(sheet) -> list[ParsedRow]:
    rows: list[ParsedRow] = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
        label = row[0]
        if label is None:
            continue
        raw_label = str(label)
        if raw_label.strip().lower() == "footnotes":
            break
        unit = normalize_unit(row[2])
        footnote_tokens = parse_footnote_tokens(row[1])
        indent = len(raw_label) - len(raw_label.lstrip(" "))
        depth = 0 if indent == 0 else max(1, indent // 5)
        raw_values = tuple(row[3:9])
        has_values = any(
            value is not None and str(value).strip().lower() not in MISSING_MARKERS for value in raw_values
        )
        is_header = not has_values and unit is None
        rows.append(
            ParsedRow(
                row_number=row_number,
                raw_label=raw_label,
                clean_label=clean_label(raw_label, footnote_tokens),
                footnote_tokens=footnote_tokens,
                unit=unit,
                depth=depth,
                is_header=is_header,
                raw_values=raw_values,
            )
        )
    return rows


def parse_workbook(path: Path) -> ParseResult:
    workbook = load_workbook(path, read_only=True, data_only=True)
    metrics: list[ParsedMetric] = []
    conflicts: list[ImportConflict] = []
    section_disclosures: dict[tuple[str, str], str] = {}

    for sheet_name in DATA_SHEETS:
        section_title = sheet_name.split(" ", 1)[1]
        sheet = workbook[sheet_name]
        rows = load_sheet_rows(sheet)
        major_group: str | None = None
        subgroup: str | None = None
        parent_labels: dict[int, str] = {}

        for index, row in enumerate(rows):
            if row.is_header:
                next_data_row = next((candidate for candidate in rows[index + 1:] if not candidate.is_header), None)
                next_data_depth = next_data_row.depth if next_data_row is not None else 0

                if row.depth == 0:
                    is_subgroup = major_group is not None and next_data_depth > 0 and row.clean_label != major_group
                    if is_subgroup:
                        subgroup = row.clean_label
                    else:
                        major_group = row.clean_label
                        subgroup = None
                    parent_labels.clear()
                    section_disclosures[(section_title, major_group)] = major_group
                else:
                    parent_labels[row.depth] = row.clean_label
                    for depth in list(parent_labels):
                        if depth > row.depth:
                            del parent_labels[depth]
                continue

            if major_group is None:
                major_group = row.clean_label
                section_disclosures[(section_title, major_group)] = major_group

            if row.depth == 0:
                if subgroup is not None:
                    subgroup = None
                parent_labels = {0: row.clean_label}
                path_components = [section_title, major_group]
                if major_group != row.clean_label:
                    path_components.append(row.clean_label)
            else:
                parent = None
                for depth in range(row.depth - 1, -1, -1):
                    if depth in parent_labels:
                        parent = parent_labels[depth]
                        break
                path_components = [section_title, major_group]
                if subgroup:
                    path_components.append(subgroup)
                if parent and parent not in {major_group, subgroup}:
                    path_components.append(parent)
                path_components.append(row.clean_label)
                parent_labels[row.depth] = row.clean_label
                for depth in list(parent_labels):
                    if depth > row.depth:
                        del parent_labels[depth]

            values_by_year: dict[int, CellValue] = {}
            value_kinds: set[str] = set()
            for year, raw_value in zip(YEARS, row.raw_values):
                parsed_value = classify_raw_value(raw_value)
                if parsed_value is None:
                    continue
                value_kinds.add(parsed_value.value_kind)
                values_by_year[year] = parsed_value

            if not values_by_year:
                continue

            tail_components = tuple(path_components[2:] if len(path_components) > 2 else path_components[1:])
            value_mode = "text" if "text" in value_kinds else "number"
            if len(value_kinds) > 1:
                conflicts.append(
                    ImportConflict(
                        code="MIXED_VALUE_REPRESENTATION",
                        severity="warning",
                        message=(
                            "Metric contains both numeric and textual cells; all imported values "
                            "for this metric will be stored as text."
                        ),
                        sheet=sheet_name,
                        row=row.row_number,
                        context={
                            "path": " / ".join(path_components),
                            "raw_values": {year: value.raw for year, value in values_by_year.items()},
                        },
                    )
                )
                coerced_values: dict[int, CellValue] = {}
                for year, value in values_by_year.items():
                    coerced_values[year] = CellValue(
                        raw=value.raw,
                        numeric_value=None,
                        text_value=value.raw,
                        value_kind="text",
                    )
                values_by_year = coerced_values

            shared_element_name = " / ".join(tail_components)
            metric = ParsedMetric(
                sheet_name=sheet_name,
                section_title=section_title,
                row_number=row.row_number,
                major_group=major_group,
                subgroup=subgroup,
                path_components=tuple(path_components),
                tail_components=tail_components,
                footnote_tokens=row.footnote_tokens,
                unit=row.unit,
                value_mode=value_mode,
                values_by_year=values_by_year,
                shared_element_code=f"BPDS_{slugify(' / '.join(path_components)).upper()}",
                shared_element_name=shared_element_name,
                requirement_item_name=shared_element_name,
                description="",
                disclosure_key=(section_title, major_group),
            )
            metric.description = build_description(metric)
            metrics.append(metric)

    metrics_by_code: dict[str, list[ParsedMetric]] = {}
    for metric in metrics:
        metrics_by_code.setdefault(metric.shared_element_code, []).append(metric)

    final_metrics: list[ParsedMetric] = []
    for code, grouped_metrics in metrics_by_code.items():
        if len(grouped_metrics) == 1:
            final_metrics.append(grouped_metrics[0])
            continue

        conflicts.append(
            ImportConflict(
                code="DUPLICATE_SHARED_ELEMENT_CODE",
                severity="warning",
                message="Detected duplicate normalized metric codes; row number suffixes were added to disambiguate them.",
                context={
                    "code": code,
                    "paths": [metric.path for metric in grouped_metrics],
                },
            )
        )
        for metric in grouped_metrics:
            metric.shared_element_code = f"{metric.shared_element_code}_R{metric.row_number}"
            final_metrics.append(metric)

    return ParseResult(
        metrics=sorted(final_metrics, key=lambda item: (item.section_title, item.row_number)),
        conflicts=conflicts,
        section_disclosures=section_disclosures,
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Import bp ESG datasheet metrics into ESGvist as proxy tenant data."
    )
    parser.add_argument("workbook", type=Path, help="Path to bp-esg-datasheet-2024.xlsx")
    parser.add_argument("--database-url", default=DEFAULT_DATABASE_URL)
    parser.add_argument("--organization-name", default=DEFAULT_ORG_NAME)
    parser.add_argument("--organization-legal-name", default=DEFAULT_ORG_LEGAL_NAME)
    parser.add_argument("--standard-code", default=DEFAULT_STANDARD_CODE)
    parser.add_argument("--standard-name", default=DEFAULT_STANDARD_NAME)
    parser.add_argument("--password", default=DEFAULT_PASSWORD)
    parser.add_argument("--report-dir", type=Path, default=DEFAULT_REPORT_DIR)
    parser.add_argument("--apply", action="store_true", help="Persist imported data.")
    return parser


async def ensure_schema(session: AsyncSession, database_url: str) -> None:
    async with session.bind.begin() as connection:
        await connection.run_sync(Base.metadata.create_all)
    await stamp_database_async(database_url)


async def get_singleton(
    session: AsyncSession,
    statement,
    *,
    conflict_code: str,
    conflict_message: str,
    conflicts: list[ImportConflict],
    context: dict[str, Any],
):
    result = (await session.execute(statement)).scalars().all()
    if len(result) > 1:
        conflicts.append(
            ImportConflict(
                code=conflict_code,
                severity="error",
                message=conflict_message,
                context=context,
            )
        )
    return result[0] if result else None


async def upsert_user(
    session: AsyncSession,
    *,
    email: str,
    full_name: str,
    password: str,
    notification_prefs: dict[str, Any] | None,
    stats: ImportStats,
) -> User:
    existing = await session.scalar(select(User).where(User.email == email))
    password_hash = hash_password(password)
    if existing is None:
        user = User(
            email=email,
            full_name=full_name,
            password_hash=password_hash,
            is_active=True,
            notification_prefs=notification_prefs,
        )
        session.add(user)
        await session.flush()
        stats.created_users += 1
        return user

    changed = False
    if existing.full_name != full_name:
        existing.full_name = full_name
        changed = True
    if not existing.is_active:
        existing.is_active = True
        changed = True
    if notification_prefs and existing.notification_prefs != notification_prefs:
        existing.notification_prefs = notification_prefs
        changed = True
    if not existing.password_hash:
        existing.password_hash = password_hash
        changed = True
    if changed:
        stats.updated_users += 1
    await session.flush()
    return existing


async def upsert_role_binding(
    session: AsyncSession,
    *,
    user_id: int,
    role: str,
    scope_type: str,
    scope_id: int | None,
    created_by: int,
    stats: ImportStats,
) -> RoleBinding:
    existing = await session.scalar(
        select(RoleBinding).where(
            RoleBinding.user_id == user_id,
            RoleBinding.scope_type == scope_type,
            RoleBinding.scope_id == scope_id,
        )
    )
    if existing is None:
        binding = RoleBinding(
            user_id=user_id,
            role=role,
            scope_type=scope_type,
            scope_id=scope_id,
            created_by=created_by,
        )
        session.add(binding)
        await session.flush()
        stats.created_role_bindings += 1
        return binding

    if existing.role != role:
        existing.role = role
        existing.created_by = created_by
        stats.updated_role_bindings += 1
    await session.flush()
    return existing


async def upsert_organization(
    session: AsyncSession,
    *,
    name: str,
    legal_name: str,
    reporting_year: int,
    stats: ImportStats,
    conflicts: list[ImportConflict],
) -> Organization:
    organization = await get_singleton(
        session,
        select(Organization).where(Organization.name == name),
        conflict_code="MULTIPLE_ORGANIZATIONS",
        conflict_message="Multiple organizations exist with the same import target name.",
        conflicts=conflicts,
        context={"organization_name": name},
    )
    if organization is None:
        organization = Organization(
            name=name,
            legal_name=legal_name,
            country="GB",
            jurisdiction="UK",
            industry="Integrated Energy",
            default_currency="USD",
            default_reporting_year=reporting_year,
            allow_password_login=True,
            allow_sso_login=True,
            enforce_sso=False,
            setup_completed=True,
            status="active",
        )
        session.add(organization)
        await session.flush()
        stats.created_organization += 1
        return organization

    changed = False
    for field_name, value in {
        "legal_name": legal_name,
        "country": "GB",
        "jurisdiction": "UK",
        "industry": "Integrated Energy",
        "default_currency": "USD",
        "default_reporting_year": reporting_year,
        "allow_password_login": True,
        "allow_sso_login": True,
        "enforce_sso": False,
        "setup_completed": True,
        "status": "active",
    }.items():
        if getattr(organization, field_name) != value:
            setattr(organization, field_name, value)
            changed = True

    if changed:
        stats.updated_organization += 1
    await session.flush()
    return organization


async def upsert_root_entity(
    session: AsyncSession,
    *,
    organization_id: int,
    legal_name: str,
    stats: ImportStats,
    conflicts: list[ImportConflict],
) -> CompanyEntity:
    entity = await get_singleton(
        session,
        select(CompanyEntity).where(
            CompanyEntity.organization_id == organization_id,
            CompanyEntity.code == "ATLAS-ROOT",
        ),
        conflict_code="MULTIPLE_ROOT_ENTITIES",
        conflict_message="Multiple root entities exist for the import organization.",
        conflicts=conflicts,
        context={"organization_id": organization_id, "code": "ATLAS-ROOT"},
    )
    if entity is None:
        entity = CompanyEntity(
            organization_id=organization_id,
            parent_entity_id=None,
            name=legal_name,
            code="ATLAS-ROOT",
            entity_type="parent_company",
            country="GB",
            jurisdiction="UK",
            status="active",
        )
        session.add(entity)
        await session.flush()
        stats.created_entities += 1
        return entity

    changed = False
    for field_name, value in {
        "name": legal_name,
        "entity_type": "parent_company",
        "country": "GB",
        "jurisdiction": "UK",
        "status": "active",
    }.items():
        if getattr(entity, field_name) != value:
            setattr(entity, field_name, value)
            changed = True
    if changed:
        stats.updated_entities += 1
    await session.flush()
    return entity


async def upsert_boundary(
    session: AsyncSession,
    *,
    organization_id: int,
    root_entity_id: int,
    stats: ImportStats,
    conflicts: list[ImportConflict],
) -> BoundaryDefinition:
    boundary = await get_singleton(
        session,
        select(BoundaryDefinition).where(
            BoundaryDefinition.organization_id == organization_id,
            BoundaryDefinition.name == "Imported ESG Datasheet Boundary",
        ),
        conflict_code="MULTIPLE_BOUNDARIES",
        conflict_message="Multiple import boundaries exist for the organization.",
        conflicts=conflicts,
        context={"organization_id": organization_id, "boundary_name": "Imported ESG Datasheet Boundary"},
    )
    if boundary is None:
        boundary = BoundaryDefinition(
            organization_id=organization_id,
            name="Imported ESG Datasheet Boundary",
            boundary_type="operational_control",
            description="Minimal reporting boundary used for the bp proxy import scenario.",
            is_default=True,
        )
        session.add(boundary)
        await session.flush()
        stats.created_boundaries += 1
    else:
        changed = False
        if boundary.boundary_type != "operational_control":
            boundary.boundary_type = "operational_control"
            changed = True
        if boundary.description != "Minimal reporting boundary used for the bp proxy import scenario.":
            boundary.description = "Minimal reporting boundary used for the bp proxy import scenario."
            changed = True
        if not boundary.is_default:
            boundary.is_default = True
            changed = True
        if changed:
            stats.updated_boundaries += 1
        await session.flush()

    membership = await session.scalar(
        select(BoundaryMembership).where(
            BoundaryMembership.boundary_definition_id == boundary.id,
            BoundaryMembership.entity_id == root_entity_id,
        )
    )
    if membership is None:
        membership = BoundaryMembership(
            boundary_definition_id=boundary.id,
            entity_id=root_entity_id,
            included=True,
            inclusion_source="manual",
            consolidation_method="full",
            inclusion_reason="Import root entity",
        )
        session.add(membership)
        await session.flush()
        stats.created_boundary_memberships += 1
    else:
        changed = False
        for field_name, value in {
            "included": True,
            "inclusion_source": "manual",
            "consolidation_method": "full",
            "inclusion_reason": "Import root entity",
        }.items():
            if getattr(membership, field_name) != value:
                setattr(membership, field_name, value)
                changed = True
        if changed:
            stats.updated_boundary_memberships += 1
        await session.flush()
    return boundary


async def upsert_standard(
    session: AsyncSession,
    *,
    code: str,
    name: str,
    stats: ImportStats,
    conflicts: list[ImportConflict],
) -> Standard:
    standard = await get_singleton(
        session,
        select(Standard).where(Standard.code == code),
        conflict_code="MULTIPLE_STANDARDS",
        conflict_message="Multiple standards exist with the same import code.",
        conflicts=conflicts,
        context={"standard_code": code},
    )
    if standard is None:
        standard = Standard(
            code=code,
            name=name,
            version="2024-proxy",
            jurisdiction="Internal",
            is_active=True,
        )
        session.add(standard)
        await session.flush()
        stats.created_standard += 1
        return standard

    changed = False
    for field_name, value in {
        "name": name,
        "version": "2024-proxy",
        "jurisdiction": "Internal",
        "is_active": True,
    }.items():
        if getattr(standard, field_name) != value:
            setattr(standard, field_name, value)
            changed = True
    if changed:
        stats.updated_standard += 1
    await session.flush()
    return standard


async def upsert_section(
    session: AsyncSession,
    *,
    standard_id: int,
    code: str,
    title: str,
    sort_order: int,
    stats: ImportStats,
) -> StandardSection:
    section = await session.scalar(
        select(StandardSection).where(
            StandardSection.standard_id == standard_id,
            StandardSection.code == code,
            StandardSection.parent_section_id.is_(None),
        )
    )
    if section is None:
        section = StandardSection(
            standard_id=standard_id,
            parent_section_id=None,
            code=code,
            title=title,
            sort_order=sort_order,
        )
        session.add(section)
        await session.flush()
        stats.created_sections += 1
        return section

    changed = False
    if section.title != title:
        section.title = title
        changed = True
    if section.sort_order != sort_order:
        section.sort_order = sort_order
        changed = True
    if changed:
        stats.updated_sections += 1
    await session.flush()
    return section


async def upsert_disclosure(
    session: AsyncSession,
    *,
    standard_id: int,
    section_id: int,
    code: str,
    title: str,
    description: str,
    sort_order: int,
    stats: ImportStats,
) -> DisclosureRequirement:
    disclosure = await session.scalar(
        select(DisclosureRequirement).where(
            DisclosureRequirement.standard_id == standard_id,
            DisclosureRequirement.code == code,
        )
    )
    if disclosure is None:
        disclosure = DisclosureRequirement(
            standard_id=standard_id,
            section_id=section_id,
            code=code,
            title=title,
            description=description,
            requirement_type="quantitative",
            mandatory_level="mandatory",
            sort_order=sort_order,
        )
        session.add(disclosure)
        await session.flush()
        stats.created_disclosures += 1
        return disclosure

    changed = False
    for field_name, value in {
        "section_id": section_id,
        "title": title,
        "description": description,
        "requirement_type": "quantitative",
        "mandatory_level": "mandatory",
        "sort_order": sort_order,
    }.items():
        if getattr(disclosure, field_name) != value:
            setattr(disclosure, field_name, value)
            changed = True
    if changed:
        stats.updated_disclosures += 1
    await session.flush()
    return disclosure


async def upsert_unit(
    session: AsyncSession,
    *,
    unit_code: str,
    category: str,
    stats: ImportStats,
) -> UnitReference:
    unit = await session.scalar(select(UnitReference).where(UnitReference.code == unit_code))
    if unit is None:
        unit = UnitReference(code=unit_code, name=unit_code, category=category)
        session.add(unit)
        await session.flush()
        stats.created_units += 1
        return unit

    changed = False
    if unit.name != unit_code:
        unit.name = unit_code
        changed = True
    if unit.category != category:
        unit.category = category
        changed = True
    if changed:
        stats.updated_units += 1
    await session.flush()
    return unit


async def upsert_shared_element(
    session: AsyncSession,
    *,
    metric: ParsedMetric,
    stats: ImportStats,
) -> SharedElement:
    element = await session.scalar(select(SharedElement).where(SharedElement.code == metric.shared_element_code))
    section_domain = SECTION_META[metric.section_title]["domain"]
    if element is None:
        element = SharedElement(
            code=metric.shared_element_code,
            name=metric.shared_element_name,
            description=metric.description,
            concept_domain=section_domain,
            default_value_type=metric.value_mode,
            default_unit_code=metric.unit,
        )
        session.add(element)
        await session.flush()
        stats.created_shared_elements += 1
        return element

    changed = False
    for field_name, value in {
        "name": metric.shared_element_name,
        "description": metric.description,
        "concept_domain": section_domain,
        "default_value_type": metric.value_mode,
        "default_unit_code": metric.unit,
        "is_current": True,
    }.items():
        if getattr(element, field_name) != value:
            setattr(element, field_name, value)
            changed = True
    if changed:
        stats.updated_shared_elements += 1
    await session.flush()
    return element


async def upsert_requirement_item(
    session: AsyncSession,
    *,
    disclosure_id: int,
    metric: ParsedMetric,
    stats: ImportStats,
) -> RequirementItem:
    item = await session.scalar(
        select(RequirementItem).where(
            RequirementItem.disclosure_requirement_id == disclosure_id,
            RequirementItem.item_code == metric.shared_element_code,
        )
    )
    if item is None:
        item = RequirementItem(
            disclosure_requirement_id=disclosure_id,
            parent_item_id=None,
            item_code=metric.shared_element_code,
            name=metric.requirement_item_name,
            description=metric.description,
            item_type="metric",
            value_type=metric.value_mode,
            unit_code=metric.unit,
            is_required=True,
            requires_evidence=False,
            cardinality_min=0,
            cardinality_max=1,
            sort_order=metric.row_number,
        )
        session.add(item)
        await session.flush()
        stats.created_requirement_items += 1
        return item

    changed = False
    for field_name, value in {
        "name": metric.requirement_item_name,
        "description": metric.description,
        "item_type": "metric",
        "value_type": metric.value_mode,
        "unit_code": metric.unit,
        "is_required": True,
        "requires_evidence": False,
        "cardinality_min": 0,
        "cardinality_max": 1,
        "sort_order": metric.row_number,
        "is_current": True,
    }.items():
        if getattr(item, field_name) != value:
            setattr(item, field_name, value)
            changed = True
    if changed:
        stats.updated_requirement_items += 1
    await session.flush()
    return item


async def ensure_mapping(
    session: AsyncSession,
    *,
    requirement_item_id: int,
    shared_element_id: int,
    stats: ImportStats,
) -> RequirementItemSharedElement:
    mapping = await session.scalar(
        select(RequirementItemSharedElement).where(
            RequirementItemSharedElement.requirement_item_id == requirement_item_id,
            RequirementItemSharedElement.shared_element_id == shared_element_id,
            RequirementItemSharedElement.is_current == True,  # noqa: E712
        )
    )
    if mapping is None:
        mapping = RequirementItemSharedElement(
            requirement_item_id=requirement_item_id,
            shared_element_id=shared_element_id,
            mapping_type="full",
            version=1,
            is_current=True,
        )
        session.add(mapping)
        await session.flush()
        stats.created_mappings += 1
    return mapping


async def upsert_project(
    session: AsyncSession,
    *,
    organization_id: int,
    boundary_definition_id: int,
    year: int,
    stats: ImportStats,
    conflicts: list[ImportConflict],
) -> ReportingProject:
    name = f"FY{year} ESG Datasheet Proxy"
    project = await get_singleton(
        session,
        select(ReportingProject).where(
            ReportingProject.organization_id == organization_id,
            ReportingProject.name == name,
        ),
        conflict_code="MULTIPLE_PROJECTS",
        conflict_message="Multiple projects exist for the same import year.",
        conflicts=conflicts,
        context={"organization_id": organization_id, "project_name": name},
    )
    desired_status = "active" if year == 2024 else "published"
    desired_deadline = date(2026, 12, 31) if year == 2024 else date(year + 1, 3, 31)

    if project is None:
        project = ReportingProject(
            organization_id=organization_id,
            name=name,
            status=desired_status,
            deadline=desired_deadline,
            reporting_year=year,
            boundary_definition_id=boundary_definition_id,
        )
        session.add(project)
        await session.flush()
        stats.created_projects += 1
        return project

    changed = False
    for field_name, value in {
        "status": desired_status,
        "deadline": desired_deadline,
        "reporting_year": year,
        "boundary_definition_id": boundary_definition_id,
    }.items():
        if getattr(project, field_name) != value:
            setattr(project, field_name, value)
            changed = True
    if changed:
        stats.updated_projects += 1
    await session.flush()
    return project


async def ensure_project_standard(
    session: AsyncSession,
    *,
    project_id: int,
    standard_id: int,
    stats: ImportStats,
) -> ReportingProjectStandard:
    existing = await session.scalar(
        select(ReportingProjectStandard).where(
            ReportingProjectStandard.reporting_project_id == project_id,
            ReportingProjectStandard.standard_id == standard_id,
        )
    )
    if existing is None:
        existing = ReportingProjectStandard(
            reporting_project_id=project_id,
            standard_id=standard_id,
            is_base_standard=True,
        )
        session.add(existing)
        await session.flush()
        stats.created_project_standards += 1
    elif not existing.is_base_standard:
        existing.is_base_standard = True
        await session.flush()
    return existing


async def upsert_assignment(
    session: AsyncSession,
    *,
    project_id: int,
    shared_element_id: int,
    entity_id: int,
    collector_id: int,
    reviewer_id: int,
    backup_collector_id: int,
    stats: ImportStats,
    conflicts: list[ImportConflict],
) -> MetricAssignment | None:
    assignments = (
        await session.execute(
            select(MetricAssignment).where(
                MetricAssignment.reporting_project_id == project_id,
                MetricAssignment.shared_element_id == shared_element_id,
                MetricAssignment.entity_id == entity_id,
                MetricAssignment.facility_id.is_(None),
            )
        )
    ).scalars().all()
    if len(assignments) > 1:
        conflicts.append(
            ImportConflict(
                code="MULTIPLE_ASSIGNMENTS",
                severity="error",
                message="Multiple assignments exist for the same project/shared element/entity key.",
                context={
                    "project_id": project_id,
                    "shared_element_id": shared_element_id,
                    "entity_id": entity_id,
                },
            )
        )
        return assignments[0]

    assignment = assignments[0] if assignments else None
    desired_deadline = date(2026, 12, 31)
    if assignment is None:
        assignment = MetricAssignment(
            reporting_project_id=project_id,
            shared_element_id=shared_element_id,
            entity_id=entity_id,
            facility_id=None,
            collector_id=collector_id,
            reviewer_id=reviewer_id,
            backup_collector_id=backup_collector_id,
            deadline=desired_deadline,
            escalation_after_days=7,
            status="completed",
        )
        session.add(assignment)
        await session.flush()
        stats.created_assignments += 1
        return assignment

    changed = False
    for field_name, value in {
        "collector_id": collector_id,
        "reviewer_id": reviewer_id,
        "backup_collector_id": backup_collector_id,
        "deadline": desired_deadline,
        "escalation_after_days": 7,
        "status": "completed",
    }.items():
        if getattr(assignment, field_name) != value:
            setattr(assignment, field_name, value)
            changed = True
    if changed:
        stats.updated_assignments += 1
    await session.flush()
    return assignment


async def upsert_data_point(
    session: AsyncSession,
    *,
    project_id: int,
    shared_element_id: int,
    entity_id: int,
    created_by: int,
    value: CellValue,
    unit_code: str | None,
    metric_value_mode: str,
    stats: ImportStats,
    conflicts: list[ImportConflict],
) -> DataPoint | None:
    data_points = (
        await session.execute(
            select(DataPoint).where(
                DataPoint.reporting_project_id == project_id,
                DataPoint.shared_element_id == shared_element_id,
                DataPoint.entity_id == entity_id,
                DataPoint.facility_id.is_(None),
            )
        )
    ).scalars().all()
    if len(data_points) > 1:
        conflicts.append(
            ImportConflict(
                code="MULTIPLE_DATA_POINTS",
                severity="error",
                message="Multiple data points exist for the same project/shared element/entity key.",
                context={
                    "project_id": project_id,
                    "shared_element_id": shared_element_id,
                    "entity_id": entity_id,
                },
            )
        )
        return data_points[0]

    desired_numeric = value.numeric_value if metric_value_mode == "number" else None
    desired_text = value.text_value if metric_value_mode == "text" else None
    if metric_value_mode == "number" and desired_numeric is None:
        desired_text = value.raw

    data_point = data_points[0] if data_points else None
    if data_point is None:
        data_point = DataPoint(
            reporting_project_id=project_id,
            shared_element_id=shared_element_id,
            entity_id=entity_id,
            facility_id=None,
            status="approved",
            numeric_value=desired_numeric,
            text_value=desired_text,
            unit_code=unit_code,
            created_by=created_by,
            review_comment="Imported from proxy ESG datasheet workbook.",
            is_derived=False,
        )
        session.add(data_point)
        await session.flush()
        stats.created_data_points += 1
        return data_point

    changed = False
    comparable_existing_numeric = (
        float(data_point.numeric_value) if data_point.numeric_value is not None else None
    )
    if comparable_existing_numeric != desired_numeric:
        data_point.numeric_value = desired_numeric
        changed = True
    if data_point.text_value != desired_text:
        data_point.text_value = desired_text
        changed = True
    if data_point.unit_code != unit_code:
        data_point.unit_code = unit_code
        changed = True
    if data_point.status != "approved":
        data_point.status = "approved"
        changed = True
    if data_point.review_comment != "Imported from proxy ESG datasheet workbook.":
        data_point.review_comment = "Imported from proxy ESG datasheet workbook."
        changed = True
    if changed:
        stats.updated_data_points += 1
    await session.flush()
    return data_point


async def ensure_item_binding(
    session: AsyncSession,
    *,
    project_id: int,
    requirement_item_id: int,
    data_point_id: int,
    stats: ImportStats,
) -> RequirementItemDataPoint:
    binding = await session.scalar(
        select(RequirementItemDataPoint).where(
            RequirementItemDataPoint.reporting_project_id == project_id,
            RequirementItemDataPoint.requirement_item_id == requirement_item_id,
            RequirementItemDataPoint.data_point_id == data_point_id,
        )
    )
    if binding is None:
        binding = RequirementItemDataPoint(
            reporting_project_id=project_id,
            requirement_item_id=requirement_item_id,
            data_point_id=data_point_id,
            binding_type="direct",
        )
        session.add(binding)
        await session.flush()
        stats.created_item_bindings += 1
    return binding


async def upsert_item_status(
    session: AsyncSession,
    *,
    project_id: int,
    requirement_item_id: int,
    status: str,
    reason: str | None,
    stats: ImportStats,
) -> RequirementItemStatus:
    current = await session.scalar(
        select(RequirementItemStatus).where(
            RequirementItemStatus.reporting_project_id == project_id,
            RequirementItemStatus.requirement_item_id == requirement_item_id,
        )
    )
    if current is None:
        current = RequirementItemStatus(
            reporting_project_id=project_id,
            requirement_item_id=requirement_item_id,
            status=status,
            status_reason=reason,
        )
        session.add(current)
        await session.flush()
        stats.created_item_statuses += 1
        return current

    if current.status != status or current.status_reason != reason:
        current.status = status
        current.status_reason = reason
        await session.flush()
        stats.updated_item_statuses += 1
    return current


async def upsert_disclosure_status(
    session: AsyncSession,
    *,
    project_id: int,
    disclosure_requirement_id: int,
    status: str,
    completion_percent: float,
    stats: ImportStats,
) -> DisclosureRequirementStatus:
    current = await session.scalar(
        select(DisclosureRequirementStatus).where(
            DisclosureRequirementStatus.reporting_project_id == project_id,
            DisclosureRequirementStatus.disclosure_requirement_id == disclosure_requirement_id,
        )
    )
    if current is None:
        current = DisclosureRequirementStatus(
            reporting_project_id=project_id,
            disclosure_requirement_id=disclosure_requirement_id,
            status=status,
            completion_percent=completion_percent,
        )
        session.add(current)
        await session.flush()
        stats.created_disclosure_statuses += 1
        return current

    if current.status != status or float(current.completion_percent) != completion_percent:
        current.status = status
        current.completion_percent = completion_percent
        await session.flush()
        stats.updated_disclosure_statuses += 1
    return current


async def upsert_evidence(
    session: AsyncSession,
    *,
    organization_id: int,
    created_by: int,
    workbook_path: Path,
    stats: ImportStats,
    conflicts: list[ImportConflict],
) -> Evidence:
    evidence = await get_singleton(
        session,
        select(Evidence).where(
            Evidence.organization_id == organization_id,
            Evidence.title == "Proxy ESG workbook source",
        ),
        conflict_code="MULTIPLE_EVIDENCE_ROWS",
        conflict_message="Multiple evidence rows exist for the imported workbook source.",
        conflicts=conflicts,
        context={"organization_id": organization_id},
    )
    description = (
        "Original workbook used as proxy source for imported ESG metrics. "
        "This evidence is linked to FY2024 imported data points."
    )
    if evidence is None:
        evidence = Evidence(
            organization_id=organization_id,
            type="file",
            title="Proxy ESG workbook source",
            description=description,
            source_type="import",
            created_by=created_by,
        )
        session.add(evidence)
        await session.flush()
        stats.created_evidences += 1
    else:
        changed = False
        if evidence.type != "file":
            evidence.type = "file"
            changed = True
        if evidence.description != description:
            evidence.description = description
            changed = True
        if evidence.source_type != "import":
            evidence.source_type = "import"
            changed = True
        if evidence.created_by != created_by:
            evidence.created_by = created_by
            changed = True
        if changed:
            stats.updated_evidences += 1
        await session.flush()

    evidence_file = await session.scalar(
        select(EvidenceFile).where(EvidenceFile.evidence_id == evidence.id)
    )
    file_uri = workbook_path.resolve().as_uri()
    if evidence_file is None:
        evidence_file = EvidenceFile(
            evidence_id=evidence.id,
            file_name=workbook_path.name,
            file_uri=file_uri,
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            file_size=workbook_path.stat().st_size,
        )
        session.add(evidence_file)
        await session.flush()
    else:
        evidence_file.file_name = workbook_path.name
        evidence_file.file_uri = file_uri
        evidence_file.mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        evidence_file.file_size = workbook_path.stat().st_size
        await session.flush()
    return evidence


async def ensure_data_point_evidence(
    session: AsyncSession,
    *,
    data_point_id: int,
    evidence_id: int,
    linked_by: int,
    stats: ImportStats,
) -> DataPointEvidence:
    link = await session.scalar(
        select(DataPointEvidence).where(
            DataPointEvidence.data_point_id == data_point_id,
            DataPointEvidence.evidence_id == evidence_id,
        )
    )
    if link is None:
        link = DataPointEvidence(
            data_point_id=data_point_id,
            evidence_id=evidence_id,
            linked_by=linked_by,
        )
        session.add(link)
        await session.flush()
        stats.created_evidence_links += 1
    return link


def build_report_markdown(summary: ImportSummary) -> str:
    lines = [
        "# bp Proxy Import Report",
        "",
        f"- Mode: `{summary.mode}`",
        f"- Workbook: `{summary.source_path}`",
        f"- Database: `{summary.database_url}`",
        f"- Organization: `{summary.organization_name}`",
        "",
        "## Counts",
        "",
        f"- Parsed metrics: `{summary.counts['parsed_metrics']}`",
        f"- Parsed data points: `{summary.counts['parsed_data_points']}`",
        f"- Metrics with 2024 values: `{summary.counts['metrics_with_2024_values']}`",
        f"- Data points by year: `{json.dumps(summary.counts['data_points_by_year'], ensure_ascii=False)}`",
        "",
        "## Projects",
        "",
    ]
    for year in YEARS:
        project_name = summary.project_names_by_year.get(year)
        project_id = summary.project_ids_by_year.get(year)
        lines.append(f"- {year}: `{project_name}` (id `{project_id}`)")

    lines.extend(
        [
            "",
            "## Credentials",
            "",
        ]
    )
    for label, email in summary.credentials.items():
        lines.append(f"- {label}: `{email}`")

    lines.append("")
    lines.append("Shared password for all seeded users:")
    lines.append("")
    lines.append("```text")
    lines.append(DEFAULT_PASSWORD if summary.mode == "dry-run" else summary.counts["password_hint"])
    lines.append("```")

    lines.extend(
        [
            "",
            "## Conflicts",
            "",
        ]
    )
    if not summary.conflicts:
        lines.append("- No conflicts detected.")
    else:
        for conflict in summary.conflicts:
            location = ""
            if conflict.get("sheet") and conflict.get("row"):
                location = f" ({conflict['sheet']} row {conflict['row']})"
            lines.append(f"- [{conflict['severity']}] `{conflict['code']}`{location}: {conflict['message']}")

    lines.extend(
        [
            "",
            "## Apply Stats",
            "",
        ]
    )
    for key, value in summary.stats.items():
        lines.append(f"- {key}: `{value}`")
    lines.append("")
    return "\n".join(lines)


def summarize_counts(parse_result: ParseResult, password: str) -> dict[str, Any]:
    data_points_by_year = {year: 0 for year in YEARS}
    for metric in parse_result.metrics:
        for year in metric.values_by_year:
            data_points_by_year[year] += 1
    return {
        "parsed_metrics": len(parse_result.metrics),
        "parsed_data_points": sum(data_points_by_year.values()),
        "metrics_with_2024_values": data_points_by_year[2024],
        "data_points_by_year": data_points_by_year,
        "password_hint": password,
    }


async def execute_import(args: argparse.Namespace) -> ImportSummary:
    parse_result = parse_workbook(args.workbook)
    stats = ImportStats()
    counts = summarize_counts(parse_result, args.password)
    report_dir = args.report_dir
    report_dir.mkdir(parents=True, exist_ok=True)

    if not args.apply:
        return ImportSummary(
            mode="dry-run",
            database_url=args.database_url,
            source_path=str(args.workbook.resolve()),
            report_dir=str(report_dir.resolve()),
            organization_name=args.organization_name,
            project_ids_by_year={},
            project_names_by_year={year: f"FY{year} ESG Datasheet Proxy" for year in YEARS},
            counts=counts,
            conflicts=[asdict(conflict) for conflict in parse_result.conflicts],
            credentials=DEFAULT_USER_EMAILS,
            stats=asdict(stats),
        )

    engine = create_async_engine(args.database_url, echo=False, pool_pre_ping=True)
    session_factory = async_sessionmaker(engine, expire_on_commit=False)

    async with session_factory() as session:
        await ensure_schema(session, args.database_url)

        admin = await upsert_user(
            session,
            email=DEFAULT_USER_EMAILS["admin"],
            full_name="Atlas Admin",
            password=args.password,
            notification_prefs={"email": True, "in_app": True},
            stats=stats,
        )
        manager = await upsert_user(
            session,
            email=DEFAULT_USER_EMAILS["manager"],
            full_name="Mila Manager",
            password=args.password,
            notification_prefs={"email": True, "in_app": True},
            stats=stats,
        )
        reviewer = await upsert_user(
            session,
            email=DEFAULT_USER_EMAILS["reviewer"],
            full_name="Roman Reviewer",
            password=args.password,
            notification_prefs={"email": True, "in_app": True},
            stats=stats,
        )
        collector = await upsert_user(
            session,
            email=DEFAULT_USER_EMAILS["collector"],
            full_name="Import Collector",
            password=args.password,
            notification_prefs={"email": True, "in_app": True},
            stats=stats,
        )
        auditor = await upsert_user(
            session,
            email=DEFAULT_USER_EMAILS["auditor"],
            full_name="Ava Auditor",
            password=args.password,
            notification_prefs={"email": True, "in_app": True},
            stats=stats,
        )

        organization = await upsert_organization(
            session,
            name=args.organization_name,
            legal_name=args.organization_legal_name,
            reporting_year=2024,
            stats=stats,
            conflicts=parse_result.conflicts,
        )
        await upsert_role_binding(
            session,
            user_id=admin.id,
            role="admin",
            scope_type="organization",
            scope_id=organization.id,
            created_by=admin.id,
            stats=stats,
        )
        await upsert_role_binding(
            session,
            user_id=manager.id,
            role="esg_manager",
            scope_type="organization",
            scope_id=organization.id,
            created_by=admin.id,
            stats=stats,
        )
        await upsert_role_binding(
            session,
            user_id=reviewer.id,
            role="reviewer",
            scope_type="organization",
            scope_id=organization.id,
            created_by=admin.id,
            stats=stats,
        )
        await upsert_role_binding(
            session,
            user_id=collector.id,
            role="collector",
            scope_type="organization",
            scope_id=organization.id,
            created_by=admin.id,
            stats=stats,
        )
        await upsert_role_binding(
            session,
            user_id=auditor.id,
            role="auditor",
            scope_type="organization",
            scope_id=organization.id,
            created_by=admin.id,
            stats=stats,
        )

        root_entity = await upsert_root_entity(
            session,
            organization_id=organization.id,
            legal_name=args.organization_legal_name,
            stats=stats,
            conflicts=parse_result.conflicts,
        )
        boundary = await upsert_boundary(
            session,
            organization_id=organization.id,
            root_entity_id=root_entity.id,
            stats=stats,
            conflicts=parse_result.conflicts,
        )

        standard = await upsert_standard(
            session,
            code=args.standard_code,
            name=args.standard_name,
            stats=stats,
            conflicts=parse_result.conflicts,
        )

        section_ids: dict[str, int] = {}
        for index, section_title in enumerate(SECTION_META.keys(), start=1):
            section = await upsert_section(
                session,
                standard_id=standard.id,
                code=SECTION_META[section_title]["code"],
                title=section_title,
                sort_order=index * 10,
                stats=stats,
            )
            section_ids[section_title] = section.id

        disclosure_ids: dict[tuple[str, str], int] = {}
        disclosure_sort_orders: dict[str, int] = {section_title: 10 for section_title in SECTION_META}
        for (section_title, major_group), title in sorted(parse_result.section_disclosures.items()):
            disclosure_code = f"{SECTION_META[section_title]['code']}-{slugify(major_group).upper()}"
            disclosure = await upsert_disclosure(
                session,
                standard_id=standard.id,
                section_id=section_ids[section_title],
                code=disclosure_code,
                title=major_group,
                description=f"Imported disclosure bucket for {section_title} / {major_group}.",
                sort_order=disclosure_sort_orders[section_title],
                stats=stats,
            )
            disclosure_ids[(section_title, major_group)] = disclosure.id
            disclosure_sort_orders[section_title] += 10

        distinct_units = sorted({metric.unit for metric in parse_result.metrics if metric.unit})
        for unit_code in distinct_units:
            category = "generic"
            for metric in parse_result.metrics:
                if metric.unit == unit_code:
                    category = SECTION_META[metric.section_title]["domain"]
                    break
            await upsert_unit(session, unit_code=unit_code, category=category, stats=stats)

        shared_element_ids: dict[str, int] = {}
        requirement_item_ids: dict[str, int] = {}
        metrics_by_code: dict[str, ParsedMetric] = {}
        metrics_by_disclosure: dict[int, list[str]] = {}
        for metric in parse_result.metrics:
            metrics_by_code[metric.shared_element_code] = metric
            shared_element = await upsert_shared_element(session, metric=metric, stats=stats)
            requirement_item = await upsert_requirement_item(
                session,
                disclosure_id=disclosure_ids[metric.disclosure_key],
                metric=metric,
                stats=stats,
            )
            await ensure_mapping(
                session,
                requirement_item_id=requirement_item.id,
                shared_element_id=shared_element.id,
                stats=stats,
            )
            shared_element_ids[metric.shared_element_code] = shared_element.id
            requirement_item_ids[metric.shared_element_code] = requirement_item.id
            metrics_by_disclosure.setdefault(disclosure_ids[metric.disclosure_key], []).append(metric.shared_element_code)

        project_ids_by_year: dict[int, int] = {}
        project_names_by_year: dict[int, str] = {}
        projects_by_year: dict[int, ReportingProject] = {}
        for year in YEARS:
            project = await upsert_project(
                session,
                organization_id=organization.id,
                boundary_definition_id=boundary.id,
                year=year,
                stats=stats,
                conflicts=parse_result.conflicts,
            )
            await ensure_project_standard(
                session,
                project_id=project.id,
                standard_id=standard.id,
                stats=stats,
            )
            projects_by_year[year] = project
            project_ids_by_year[year] = project.id
            project_names_by_year[year] = project.name

        evidence = await upsert_evidence(
            session,
            organization_id=organization.id,
            created_by=admin.id,
            workbook_path=args.workbook,
            stats=stats,
            conflicts=parse_result.conflicts,
        )

        for year, project in projects_by_year.items():
            disclosure_applicable_counts: dict[int, dict[str, int]] = {
                disclosure_id: {"complete": 0, "applicable": 0}
                for disclosure_id in metrics_by_disclosure
            }
            for metric in parse_result.metrics:
                requirement_item_id = requirement_item_ids[metric.shared_element_code]
                disclosure_id = disclosure_ids[metric.disclosure_key]
                cell_value = metric.values_by_year.get(year)
                if cell_value is None:
                    await upsert_item_status(
                        session,
                        project_id=project.id,
                        requirement_item_id=requirement_item_id,
                        status="not_applicable",
                        reason="No source value in workbook for this reporting year.",
                        stats=stats,
                    )
                    continue

                data_point = await upsert_data_point(
                    session,
                    project_id=project.id,
                    shared_element_id=shared_element_ids[metric.shared_element_code],
                    entity_id=root_entity.id,
                    created_by=collector.id,
                    value=cell_value,
                    unit_code=metric.unit,
                    metric_value_mode=metric.value_mode,
                    stats=stats,
                    conflicts=parse_result.conflicts,
                )
                if data_point is None:
                    continue

                await ensure_item_binding(
                    session,
                    project_id=project.id,
                    requirement_item_id=requirement_item_id,
                    data_point_id=data_point.id,
                    stats=stats,
                )
                await upsert_item_status(
                    session,
                    project_id=project.id,
                    requirement_item_id=requirement_item_id,
                    status="complete",
                    reason="Imported from workbook.",
                    stats=stats,
                )
                disclosure_applicable_counts[disclosure_id]["complete"] += 1
                disclosure_applicable_counts[disclosure_id]["applicable"] += 1

                if year == 2024:
                    await upsert_assignment(
                        session,
                        project_id=project.id,
                        shared_element_id=shared_element_ids[metric.shared_element_code],
                        entity_id=root_entity.id,
                        collector_id=collector.id,
                        reviewer_id=reviewer.id,
                        backup_collector_id=manager.id,
                        stats=stats,
                        conflicts=parse_result.conflicts,
                    )
                    await ensure_data_point_evidence(
                        session,
                        data_point_id=data_point.id,
                        evidence_id=evidence.id,
                        linked_by=admin.id,
                        stats=stats,
                    )

            for disclosure_id, buckets in disclosure_applicable_counts.items():
                applicable = buckets["applicable"]
                complete = buckets["complete"]
                if applicable == 0:
                    status = "missing"
                    completion_percent = 0.0
                else:
                    completion_percent = round((complete / applicable) * 100, 1)
                    status = "complete" if complete == applicable else "partial"
                await upsert_disclosure_status(
                    session,
                    project_id=project.id,
                    disclosure_requirement_id=disclosure_id,
                    status=status,
                    completion_percent=completion_percent,
                    stats=stats,
                )

        await session.commit()

    await engine.dispose()
    return ImportSummary(
        mode="apply",
        database_url=args.database_url,
        source_path=str(args.workbook.resolve()),
        report_dir=str(report_dir.resolve()),
        organization_name=args.organization_name,
        project_ids_by_year=project_ids_by_year,
        project_names_by_year=project_names_by_year,
        counts=counts,
        conflicts=[asdict(conflict) for conflict in parse_result.conflicts],
        credentials={
            "admin": admin.email,
            "manager": manager.email,
            "reviewer": reviewer.email,
            "collector": collector.email,
            "auditor": auditor.email,
        },
        stats=asdict(stats),
    )


def write_summary_files(summary: ImportSummary) -> None:
    report_dir = Path(summary.report_dir)
    report_dir.mkdir(parents=True, exist_ok=True)

    (report_dir / "import-summary.json").write_text(
        json.dumps(asdict(summary), indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    (report_dir / "import-report.md").write_text(
        build_report_markdown(summary),
        encoding="utf-8",
    )
    credentials_lines = [
        "# Import Credentials",
        "",
        f"- Organization: `{summary.organization_name}`",
        "",
    ]
    for label, email in summary.credentials.items():
        credentials_lines.append(f"- {label}: `{email}`")
    credentials_lines.extend(
        [
            "",
            "Shared password:",
            "",
            "```text",
            summary.counts["password_hint"],
            "```",
            "",
        ]
    )
    (report_dir / "credentials.md").write_text("\n".join(credentials_lines), encoding="utf-8")


async def main() -> None:
    args = build_parser().parse_args()
    summary = await execute_import(args)
    write_summary_files(summary)
    print(json.dumps(asdict(summary), indent=2, ensure_ascii=False))


if __name__ == "__main__":
    asyncio.run(main())
