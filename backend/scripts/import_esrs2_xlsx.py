from __future__ import annotations

import argparse
import asyncio
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

BACKEND_DIR = Path(__file__).resolve().parents[1]
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from app.core.config import settings
from scripts.import_esrs_xlsx import (
    EsrsBlock,
    EsrsRow,
    EsrsSubsection,
    build_disclosure_description,
    build_applicability_rule,
    build_parsed_data_points,
    get_or_create_standard,
    infer_disclosure_type,
    normalize_bullets,
    normalize_text,
    upsert_esrs_requirement_item,
    upsert_section,
)
from scripts.import_gri_docx import (
    ImportStats,
    deactivate_stale_requirement_items,
    ensure_mapping,
    upsert_shared_element,
)


STANDARD_CODE = "ESRS 2"
STANDARD_NAME = "ESRS 2: General Disclosures"
ROW_REF_RE = re.compile(r"^(?P<disclosure_code>(?:BP|GOV|SBM|IRO)-\d+|GDR-[A-Z])(?:\s+(?P<clause_ref>.+))?$")

SECTION_TITLES: dict[str, str] = {
    "BP": "Basis for Preparation",
    "GOV": "Governance",
    "SBM": "Strategy, Business Model and Value Chain",
    "IRO": "Impact, Risk and Opportunity Management",
    "GDR": "General Disclosure Requirements",
}

DISCLOSURE_TITLES: dict[str, str] = {
    "BP-1": "General Basis for Preparation of Sustainability Statements",
    "BP-2": "Disclosures in Relation to Specific Circumstances",
    "GOV-1": "Role of the Administrative, Management and Supervisory Bodies",
    "GOV-2": "Information Provided to and Sustainability Matters Addressed by the Administrative, Management and Supervisory Bodies",
    "GOV-3": "Integration of Sustainability-Related Performance in Incentive Schemes",
    "GOV-4": "Statement on Due Diligence",
    "SBM-1": "Strategy, Business Model and Value Chain",
    "SBM-2": "Interests and Views of Stakeholders",
    "SBM-3": "Material Impacts, Risks and Opportunities and Their Interaction with Strategy and Business Model",
    "IRO-1": "Description of the Processes to Identify and Assess Material Impacts, Risks and Opportunities",
    "IRO-2": "Disclosure Requirements in ESRS Covered by the Undertaking's Sustainability Statement",
    "GDR-P": "Policies Adopted to Manage Material Sustainability Matters",
    "GDR-A": "Actions and Resources in Relation to Material Sustainability Matters",
    "GDR-M": "Metrics in Relation to Material Sustainability Matters",
    "GDR-T": "Targets in Relation to Material Sustainability Matters",
}


@dataclass
class Esrs2Section:
    code: str
    title: str
    sort_order: int
    disclosures: list[EsrsBlock] = field(default_factory=list)


@dataclass
class Esrs2ParsedWorkbook:
    standard_code: str
    standard_name: str
    standard_version: str
    standard_jurisdiction: str
    sections: list[Esrs2Section]


def build_standard_payload(parsed: Esrs2ParsedWorkbook):
    return SimpleNamespace(
        standard_code=parsed.standard_code,
        standard_name=parsed.standard_name,
        standard_version=parsed.standard_version,
        standard_jurisdiction=parsed.standard_jurisdiction,
    )


async def upsert_disclosure_esrs2(
    session: AsyncSession,
    *,
    standard_id: int,
    section_id: int,
    section: Esrs2Section,
    block: EsrsBlock,
    stats: ImportStats,
):
    from sqlalchemy import select

    from app.db.models.standard import DisclosureRequirement

    disclosure = (
        await session.execute(
            select(DisclosureRequirement).where(
                DisclosureRequirement.standard_id == standard_id,
                DisclosureRequirement.code == block.disclosure_code,
            )
        )
    ).scalar_one_or_none()

    parsed_items = [parsed for _row, parsed in build_parsed_data_points(block)]
    description = build_disclosure_description(block)
    applicability_rule = build_applicability_rule(block)
    applicability_rule["source_format"] = "esrs2_xlsx"
    applicability_rule["section_code"] = section.code
    applicability_rule["section_title"] = section.title
    title = block.title
    requirement_type = infer_disclosure_type(parsed_items)

    if disclosure is None:
        disclosure = DisclosureRequirement(
            standard_id=standard_id,
            section_id=section_id,
            code=block.disclosure_code,
            title=title,
            description=description,
            requirement_type=requirement_type,
            mandatory_level="mandatory",
            applicability_rule=applicability_rule,
            sort_order=block.sort_order,
        )
        session.add(disclosure)
        await session.flush()
        stats.created_disclosures += 1
        return disclosure

    changed = False
    for field_name, value in {
        "section_id": section_id,
        "title": title,
        "description": description,
        "requirement_type": requirement_type,
        "mandatory_level": "mandatory",
        "applicability_rule": applicability_rule,
        "sort_order": block.sort_order,
    }.items():
        if getattr(disclosure, field_name) != value:
            setattr(disclosure, field_name, value)
            changed = True
    if changed:
        await session.flush()
        stats.updated_disclosures += 1
    return disclosure


def disclosure_title_for(code: str, first_row_source: str | None) -> str:
    if code in DISCLOSURE_TITLES:
        return DISCLOSURE_TITLES[code]
    if first_row_source:
        return first_row_source
    return code


def parse_workbook(path: Path) -> Esrs2ParsedWorkbook:
    workbook = load_workbook(path, read_only=True, data_only=True)

    sections_by_code: dict[str, Esrs2Section] = {}
    section_order: list[str] = []
    for sheet in workbook.worksheets:
        current_sheet_disclosures: dict[str, EsrsBlock] = {}

        for row in sheet.iter_rows(values_only=True):
            cells = ["" if cell is None else str(cell).strip() for cell in row[:8]]
            ref = normalize_text(cells[0])
            source = normalize_text(cells[1])
            standardised_requirement = normalize_text(cells[2])
            interpretation = normalize_text(cells[3]) or None
            data_points = normalize_bullets(cells[4])
            evidence = normalize_text(cells[5]) or None
            owner = normalize_text(cells[6]) or None
            frequency = normalize_text(cells[7]) or None

            if not any((ref, source, standardised_requirement, interpretation, data_points, evidence, owner, frequency)):
                continue
            if ref.lower() == "ref" or ref.startswith("BLOCK "):
                continue

            row_match = ROW_REF_RE.match(ref)
            if row_match is None:
                raise ValueError(f"Unrecognized ESRS 2 row reference {ref!r} in sheet {sheet.title!r}")

            disclosure_code = row_match.group("disclosure_code")
            clause_ref = normalize_text(row_match.group("clause_ref")) or disclosure_code
            section_code = disclosure_code.split("-", 1)[0]

            if section_code not in sections_by_code:
                sections_by_code[section_code] = Esrs2Section(
                    code=section_code,
                    title=SECTION_TITLES.get(section_code, section_code),
                    sort_order=len(section_order) + 1,
                )
                section_order.append(section_code)

            if disclosure_code not in current_sheet_disclosures:
                block = EsrsBlock(
                    sort_order=len(sections_by_code[section_code].disclosures) + 1,
                    disclosure_code=disclosure_code,
                    title=disclosure_title_for(disclosure_code, source or None),
                    subsections=[EsrsSubsection(title="Requirements", row_kind="requirements")],
                )
                current_sheet_disclosures[disclosure_code] = block
                sections_by_code[section_code].disclosures.append(block)

            block = current_sheet_disclosures[disclosure_code]
            block.subsections[0].rows.append(
                EsrsRow(
                    ref=ref,
                    clause_ref=clause_ref,
                    source=source,
                    standardised_requirement=standardised_requirement,
                    interpretation=interpretation,
                    data_points=data_points,
                    evidence=evidence,
                    owner=owner,
                    frequency=frequency,
                    row_kind="requirements",
                )
            )

    sections = [sections_by_code[code] for code in section_order]
    if not sections:
        raise ValueError(f"No ESRS 2 disclosures found in workbook {path}")

    return Esrs2ParsedWorkbook(
        standard_code=STANDARD_CODE,
        standard_name=STANDARD_NAME,
        standard_version="2024",
        standard_jurisdiction="EU",
        sections=sections,
    )


async def import_workbook(
    session: AsyncSession,
    parsed: Esrs2ParsedWorkbook,
    *,
    with_shared_elements: bool,
) -> tuple[ImportStats, int]:
    stats = ImportStats()
    standard = await get_or_create_standard(session, build_standard_payload(parsed))
    total_items = 0

    for section in parsed.sections:
        section_row = await upsert_section(
            session,
            standard_id=standard.id,
            code=section.code,
            title=section.title,
            sort_order=section.sort_order,
            parent_section_id=None,
            stats=stats,
        )

        for disclosure_block in section.disclosures:
            disclosure = await upsert_disclosure_esrs2(
                session,
                standard_id=standard.id,
                section_id=section_row.id,
                section=section,
                block=disclosure_block,
                stats=stats,
            )

            active_item_codes: set[str] = set()
            for sort_order, (row, datapoint) in enumerate(build_parsed_data_points(disclosure_block), start=1):
                item = await upsert_esrs_requirement_item(
                    session,
                    disclosure_id=disclosure.id,
                    datapoint=datapoint,
                    sort_order=sort_order,
                    description=datapoint.label.split(" - ", 1)[-1],
                    metadata={
                        "source_ref": row.ref,
                        "clause_ref": row.clause_ref,
                        "row_kind": row.row_kind,
                        "source": row.source,
                        "interpretation": row.interpretation,
                        "evidence": row.evidence,
                        "owner": row.owner,
                        "frequency": row.frequency,
                        "section_code": section.code,
                        "section_title": section.title,
                    },
                    stats=stats,
                )
                active_item_codes.add(datapoint.item_code)
                total_items += 1
                if with_shared_elements:
                    concept_domain = re.sub(
                        r"[^a-z0-9]+",
                        "_",
                        f"{parsed.standard_code}_{section.code}_{disclosure_block.title}".lower(),
                    ).strip("_")
                    element = await upsert_shared_element(session, item, concept_domain, stats)
                    await ensure_mapping(session, item.id, element.id, stats)

            await deactivate_stale_requirement_items(session, disclosure.id, active_item_codes)

    return stats, total_items


async def run_single_import(
    session_factory: async_sessionmaker[AsyncSession],
    path: Path,
    *,
    apply: bool,
    with_shared_elements: bool,
) -> tuple[Esrs2ParsedWorkbook, ImportStats, int]:
    parsed = parse_workbook(path)
    async with session_factory() as session:
        stats, total_items = await import_workbook(session, parsed, with_shared_elements=with_shared_elements)
        if apply:
            await session.commit()
        else:
            await session.rollback()
    return parsed, stats, total_items


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Import ESRS 2 XLSX workbook into the framework catalog.")
    parser.add_argument(
        "xlsx_path",
        type=Path,
        help="Path to ESRS2.xlsx.",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Persist changes. Without this flag, the script runs as a dry-run.",
    )
    parser.add_argument(
        "--no-shared-elements",
        action="store_true",
        help="Skip creating shared elements and mappings.",
    )
    return parser


async def async_main(args: argparse.Namespace) -> int:
    engine = create_async_engine(settings.database_url, echo=False)
    session_factory = async_sessionmaker(engine, expire_on_commit=False)

    try:
        parsed, stats, total_items = await run_single_import(
            session_factory,
            args.xlsx_path,
            apply=args.apply,
            with_shared_elements=not args.no_shared_elements,
        )
    finally:
        await engine.dispose()

    disclosure_count = sum(len(section.disclosures) for section in parsed.sections)
    print(f"Input: {args.xlsx_path}")
    print(f"Mode: {'apply' if args.apply else 'dry-run'}")
    print(f"Standard: {parsed.standard_code} - {parsed.standard_name}")
    print(f"Sections: {len(parsed.sections)}")
    print(f"Disclosures: {disclosure_count}")
    print(f"Datapoints: {total_items}")
    print(f"Sections:         +{stats.created_sections} new, {stats.updated_sections} updated")
    print(f"Disclosures:      +{stats.created_disclosures} new, {stats.updated_disclosures} updated")
    print(f"RequirementItems: +{stats.created_items} new, {stats.updated_items} updated")
    print(f"SharedElements:   +{stats.created_shared_elements} new, {stats.updated_shared_elements} updated")
    print(f"Mappings:         +{stats.created_mappings} new")
    return 0


def main() -> int:
    parser = build_arg_parser()
    args = parser.parse_args()
    return asyncio.run(async_main(args))


if __name__ == "__main__":
    raise SystemExit(main())
