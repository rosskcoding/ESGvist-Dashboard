from __future__ import annotations

import argparse
import asyncio
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

BACKEND_DIR = Path(__file__).resolve().parents[1]
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from app.core.config import settings
from app.db.models.requirement_item import RequirementItem
from app.db.models.standard import DisclosureRequirement, Standard, StandardSection
from scripts.import_gri_docx import (
    ImportStats,
    ParsedDataPoint,
    deactivate_stale_requirement_items,
    ensure_mapping,
    infer_disclosure_type,
    upsert_shared_element,
)

BLOCK_HEADER_RE = re.compile(
    r"^🔹\s*BLOCK\s+(?P<sort>\d+)\s+[—-]\s+(?P<code>[A-Z]\d+-\d+)\s+(?P<title>.+)$"
)
SUBSECTION_RE = re.compile(r"^🔸\s*(?P<title>.+?)\s*$")
ROW_REF_RE = re.compile(r"^(?P<disclosure_code>[A-Z]\d+-\d+)\s+(?P<clause_ref>§[\w().-]+|AR\d+)$")

STANDARD_NAME_BY_STEM: dict[str, tuple[str, str]] = {
    "E1": ("ESRS E1", "ESRS E1: Climate Change"),
    "E2": ("ESRS E2", "ESRS E2: Pollution"),
    "E3": ("ESRS E3", "ESRS E3: Water and Marine Resources"),
    "E4": ("ESRS E4", "ESRS E4: Biodiversity and Ecosystems"),
    "E5": ("ESRS E5", "ESRS E5: Resource Use and Circular Economy"),
    "G1": ("ESRS G1", "ESRS G1: Business Conduct"),
    "S1": ("ESRS S1", "ESRS S1: Own Workforce"),
    "S2": ("ESRS S2", "ESRS S2: Workers in the Value Chain"),
    "S3": ("ESRS S3", "ESRS S3: Affected Communities"),
    "S4": ("ESRS S4", "ESRS S4: Consumers and End-users"),
}


@dataclass
class EsrsRow:
    ref: str
    clause_ref: str
    source: str
    standardised_requirement: str
    interpretation: str | None
    data_points: list[str]
    evidence: str | None
    owner: str | None
    frequency: str | None
    row_kind: str


@dataclass
class EsrsSubsection:
    title: str
    row_kind: str
    rows: list[EsrsRow] = field(default_factory=list)


@dataclass
class EsrsBlock:
    sort_order: int
    disclosure_code: str
    title: str
    subsections: list[EsrsSubsection] = field(default_factory=list)


@dataclass
class EsrsParsedWorkbook:
    standard_code: str
    standard_name: str
    standard_version: str
    standard_jurisdiction: str
    blocks: list[EsrsBlock]


def normalize_text(value: str | None) -> str:
    return re.sub(r"\s+", " ", (value or "").strip())


def slugify(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")


def normalize_bullets(raw_value: str | None) -> list[str]:
    if raw_value is None:
        return []
    lines = []
    for part in str(raw_value).replace("\r", "\n").split("\n"):
        cleaned = normalize_text(part).lstrip("-").strip()
        if cleaned and cleaned != "#VALUE!":
            lines.append(cleaned)
    return lines


def infer_item_shape(label: str) -> tuple[str, str, str | None]:
    normalized = normalize_text(label).lower()
    if "(y/n)" in normalized or normalized.endswith("(y/n)"):
        return "attribute", "boolean", None
    if normalized.endswith("(m³)") or " (m³)" in normalized:
        return "metric", "number", "m3"
    if normalized.endswith("(%)") or " (%)" in normalized or normalized.endswith("%"):
        return "metric", "number", "%"
    if normalized.endswith("(aed or usd)") or normalized.endswith("(usd/aed)"):
        return "metric", "number", "currency"
    if normalized.startswith(("number ", "total ", "average ", "coverage ", "capex", "opex", "turnover rate")):
        return "metric", "number", None
    if normalized.startswith(("timeline", "start date", "end date")):
        return "attribute", "text", None
    if normalized.startswith(("policy ", "scope ", "coverage ", "objectives", "standards ", "indicator ", "threshold ", "methodology", "units ", "groups ", "regions", "countries", "site", "actions ", "resources ", "target", "baseline", "metric ", "channels ", "agreements", "works council", "family leave")):
        return "attribute", "text", None
    if normalized.startswith(("forced labour", "child labour", "trafficking", "decision influence")):
        return "attribute", "boolean", None
    return "attribute", "text", None


def standard_from_path(path: Path) -> tuple[str, str]:
    stem = path.stem.strip().upper()
    if stem in STANDARD_NAME_BY_STEM:
        return STANDARD_NAME_BY_STEM[stem]
    return (f"ESRS {stem}", f"ESRS {stem}")


def subsection_kind_from_title(title: str) -> str:
    normalized = normalize_text(title).lower()
    if normalized == "application requirements":
        return "application_requirement"
    if normalized == "core metrics":
        return "core_metrics"
    return "requirements"


def find_or_create_subsection(block: EsrsBlock, title: str, row_kind: str) -> EsrsSubsection:
    for subsection in block.subsections:
        if subsection.title == title and subsection.row_kind == row_kind:
            return subsection
    subsection = EsrsSubsection(title=title, row_kind=row_kind)
    block.subsections.append(subsection)
    return subsection


def parse_workbook(path: Path) -> EsrsParsedWorkbook:
    if path.stem.strip().upper() == "ESRS2":
        raise ValueError("ESRS2.xlsx must be imported via import_esrs2_xlsx.py")
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    standard_code, standard_name = standard_from_path(path)

    blocks: list[EsrsBlock] = []
    current_block: EsrsBlock | None = None
    current_subsection: EsrsSubsection | None = None
    current_row: EsrsRow | None = None

    for row in sheet.iter_rows(values_only=True):
        cells = ["" if cell is None else str(cell).strip() for cell in row[:8]]
        ref = normalize_text(cells[0])
        source = normalize_text(cells[1])
        standardised_requirement = normalize_text(cells[2])
        interpretation = normalize_text(cells[3]) or None
        data_points = normalize_bullets(cells[4])
        evidence = normalize_text(cells[5]) or None
        owner = normalize_text(cells[6]) or None
        frequency = normalize_text(cells[7]) or None

        if not any((ref, source, standardised_requirement, interpretation, data_points, evidence, owner, frequency)):
            current_row = None
            continue

        block_match = BLOCK_HEADER_RE.match(ref)
        if block_match:
            current_block = EsrsBlock(
                sort_order=int(block_match.group("sort")),
                disclosure_code=block_match.group("code"),
                title=normalize_text(block_match.group("title")),
            )
            blocks.append(current_block)
            current_subsection = find_or_create_subsection(current_block, "Requirements", "requirements")
            current_row = None
            continue

        subsection_match = SUBSECTION_RE.match(ref)
        if subsection_match and current_block is not None:
            subsection_title = normalize_text(subsection_match.group("title"))
            current_subsection = find_or_create_subsection(
                current_block,
                subsection_title,
                subsection_kind_from_title(subsection_title),
            )
            current_row = None
            continue

        if ref.lower() == "ref":
            current_row = None
            continue

        row_match = ROW_REF_RE.match(ref)
        if row_match and current_block is not None:
            if row_match.group("disclosure_code") != current_block.disclosure_code:
                raise ValueError(
                    f"Row {ref!r} does not match active block {current_block.disclosure_code!r} in {path.name}"
                )
            if current_subsection is None:
                current_subsection = find_or_create_subsection(current_block, "Requirements", "requirements")

            current_row = EsrsRow(
                ref=ref,
                clause_ref=row_match.group("clause_ref"),
                source=source,
                standardised_requirement=standardised_requirement,
                interpretation=interpretation,
                data_points=data_points,
                evidence=evidence,
                owner=owner,
                frequency=frequency,
                row_kind=current_subsection.row_kind,
            )
            current_subsection.rows.append(current_row)
            continue

        if current_row is not None and data_points:
            current_row.data_points.extend(data_points)

    if not blocks:
        raise ValueError(f"No ESRS blocks found in workbook {path}")

    return EsrsParsedWorkbook(
        standard_code=standard_code,
        standard_name=standard_name,
        standard_version="2024",
        standard_jurisdiction="EU",
        blocks=blocks,
    )


def clean_data_points(data_points: Iterable[str]) -> list[str]:
    cleaned: list[str] = []
    seen: set[str] = set()
    for data_point in data_points:
        normalized = normalize_text(data_point)
        if not normalized or normalized == "#VALUE!":
            continue
        if normalized not in seen:
            seen.add(normalized)
            cleaned.append(normalized)
    return cleaned


def build_requirement_item_name(clause_ref: str, label: str) -> str:
    return f"{clause_ref} - {label}"


async def upsert_esrs_requirement_item(
    session: AsyncSession,
    *,
    disclosure_id: int,
    datapoint: ParsedDataPoint,
    sort_order: int,
    description: str,
    metadata: dict,
    stats: ImportStats,
) -> RequirementItem:
    item = (
        await session.execute(
            select(RequirementItem).where(
                RequirementItem.disclosure_requirement_id == disclosure_id,
                RequirementItem.item_code == datapoint.item_code,
                RequirementItem.is_current.is_(True),
            )
        )
    ).scalar_one_or_none()

    if item is None:
        item = RequirementItem(
            disclosure_requirement_id=disclosure_id,
            item_code=datapoint.item_code,
            name=datapoint.label,
            description=description,
            item_type=datapoint.item_type,
            value_type=datapoint.value_type,
            unit_code=datapoint.unit_code,
            is_required=True,
            requires_evidence=False,
            cardinality_min=0,
            cardinality_max=None,
            sort_order=sort_order,
            granularity_rule=metadata,
        )
        session.add(item)
        await session.flush()
        stats.created_items += 1
        return item

    changed = False
    updates = {
        "name": datapoint.label,
        "description": description,
        "item_type": datapoint.item_type,
        "value_type": datapoint.value_type,
        "unit_code": datapoint.unit_code,
        "sort_order": sort_order,
        "granularity_rule": metadata,
    }
    for field_name, value in updates.items():
        if getattr(item, field_name) != value:
            setattr(item, field_name, value)
            changed = True
    if changed:
        await session.flush()
        stats.updated_items += 1
    return item


def build_content_block(subsection: EsrsSubsection) -> dict:
    row_items = [f"{row.ref} - {row.standardised_requirement}" for row in subsection.rows]
    body_lines = []
    for row in subsection.rows:
        parts = [f"{row.ref}: {row.standardised_requirement}"]
        if row.interpretation:
            parts.append(f"Interpretation: {row.interpretation}")
        if row.evidence:
            parts.append(f"Evidence: {row.evidence}")
        body_lines.append("\n".join(parts))

    return {
        "type": subsection.row_kind,
        "title": "Application requirements" if subsection.row_kind == "application_requirement" else subsection.title,
        "items": row_items,
        "paragraphs": [],
        "body_md": "\n\n".join(body_lines),
        "metadata": {
            "rows": [
                {
                    "ref": row.ref,
                    "clause_ref": row.clause_ref,
                    "source": row.source,
                    "standardised_requirement": row.standardised_requirement,
                    "interpretation": row.interpretation,
                    "data_points": row.data_points,
                    "evidence": row.evidence,
                    "owner": row.owner,
                    "frequency": row.frequency,
                    "row_kind": row.row_kind,
                }
                for row in subsection.rows
            ]
        },
    }


def build_disclosure_description(block: EsrsBlock) -> str:
    sections: list[str] = []
    for subsection in block.subsections:
        title = "Application requirements" if subsection.row_kind == "application_requirement" else subsection.title
        lines = [title]
        for row in subsection.rows:
            lines.append(f"{row.ref}: {row.standardised_requirement}")
            if row.interpretation:
                lines.append(f"Interpretation: {row.interpretation}")
            if row.evidence:
                lines.append(f"Evidence: {row.evidence}")
        sections.append("\n".join(lines))
    return "\n\n".join(sections).strip()


def build_applicability_rule(block: EsrsBlock) -> dict:
    return {
        "source_format": "esrs_xlsx",
        "disclosure_code": block.disclosure_code,
        "display_term_requirement": "Standardised Requirement",
        "content_blocks": [build_content_block(subsection) for subsection in block.subsections if subsection.rows],
    }


def build_parsed_data_points(block: EsrsBlock) -> list[tuple[EsrsRow, ParsedDataPoint]]:
    clause_occurrence_totals: dict[str, int] = {}
    ordered_rows: list[EsrsRow] = []
    for subsection in block.subsections:
        for row in subsection.rows:
            ordered_rows.append(row)
            clause_occurrence_totals[row.clause_ref] = clause_occurrence_totals.get(row.clause_ref, 0) + 1

    clause_occurrence_seen: dict[str, int] = {}
    parsed: list[tuple[EsrsRow, ParsedDataPoint]] = []
    for row in ordered_rows:
        clause_occurrence_seen[row.clause_ref] = clause_occurrence_seen.get(row.clause_ref, 0) + 1
        occurrence_index = clause_occurrence_seen[row.clause_ref]
        clause_slug = slugify(row.clause_ref)
        clause_has_repeats = clause_occurrence_totals[row.clause_ref] > 1

        for item_index, label in enumerate(clean_data_points(row.data_points), start=1):
            item_type, value_type, unit_code = infer_item_shape(label)
            if clause_has_repeats:
                item_code = f"{block.disclosure_code}.{clause_slug}.r{occurrence_index}.{item_index}"
            else:
                item_code = f"{block.disclosure_code}.{clause_slug}.{item_index}"
            parsed.append(
                (
                    row,
                    ParsedDataPoint(
                        raw_code=row.ref,
                        item_code=item_code,
                        label=build_requirement_item_name(row.clause_ref, label),
                        item_type=item_type,
                        value_type=value_type,
                        unit_code=unit_code,
                    ),
                )
            )
    return parsed


async def get_or_create_standard(session: AsyncSession, parsed: EsrsParsedWorkbook) -> Standard:
    standard = (
        await session.execute(select(Standard).where(Standard.code == parsed.standard_code))
    ).scalar_one_or_none()
    if standard is None:
        standard = Standard(
            code=parsed.standard_code,
            name=parsed.standard_name,
            version=parsed.standard_version,
            jurisdiction=parsed.standard_jurisdiction,
            is_active=True,
        )
        session.add(standard)
        await session.flush()
        return standard

    changed = False
    for field_name, value in {
        "name": parsed.standard_name,
        "version": parsed.standard_version,
        "jurisdiction": parsed.standard_jurisdiction,
        "is_active": True,
    }.items():
        if getattr(standard, field_name) != value:
            setattr(standard, field_name, value)
            changed = True
    if changed:
        await session.flush()
    return standard


async def upsert_section(
    session: AsyncSession,
    *,
    standard_id: int,
    code: str,
    title: str,
    sort_order: int,
    parent_section_id: int | None,
    stats: ImportStats,
) -> StandardSection:
    section = (
        await session.execute(
            select(StandardSection).where(
                StandardSection.standard_id == standard_id,
                StandardSection.code == code,
            )
        )
    ).scalar_one_or_none()

    if section is None:
        section = StandardSection(
            standard_id=standard_id,
            parent_section_id=parent_section_id,
            code=code,
            title=title,
            sort_order=sort_order,
        )
        session.add(section)
        await session.flush()
        stats.created_sections += 1
        return section

    changed = False
    for field_name, value in {
        "parent_section_id": parent_section_id,
        "title": title,
        "sort_order": sort_order,
    }.items():
        if getattr(section, field_name) != value:
            setattr(section, field_name, value)
            changed = True
    if changed:
        await session.flush()
        stats.updated_sections += 1
    return section


async def upsert_disclosure(
    session: AsyncSession,
    *,
    standard_id: int,
    section_id: int,
    block: EsrsBlock,
    stats: ImportStats,
) -> DisclosureRequirement:
    disclosure = (
        await session.execute(
            select(DisclosureRequirement).where(
                DisclosureRequirement.standard_id == standard_id,
                DisclosureRequirement.code == block.disclosure_code,
            )
        )
    ).scalar_one_or_none()

    parsed_items = [parsed for _row, parsed in build_parsed_data_points(block)]
    description = build_disclosure_description(block)
    applicability_rule = build_applicability_rule(block)
    title = block.title
    requirement_type = infer_disclosure_type(parsed_items)

    if disclosure is None:
        disclosure = DisclosureRequirement(
            standard_id=standard_id,
            section_id=section_id,
            code=block.disclosure_code,
            title=title,
            description=description,
            requirement_type=requirement_type,
            mandatory_level="mandatory",
            applicability_rule=applicability_rule,
            sort_order=block.sort_order,
        )
        session.add(disclosure)
        await session.flush()
        stats.created_disclosures += 1
        return disclosure

    changed = False
    for field_name, value in {
        "section_id": section_id,
        "title": title,
        "description": description,
        "requirement_type": requirement_type,
        "mandatory_level": "mandatory",
        "applicability_rule": applicability_rule,
        "sort_order": block.sort_order,
    }.items():
        if getattr(disclosure, field_name) != value:
            setattr(disclosure, field_name, value)
            changed = True
    if changed:
        await session.flush()
        stats.updated_disclosures += 1
    return disclosure


async def import_workbook(
    session: AsyncSession,
    parsed: EsrsParsedWorkbook,
    *,
    with_shared_elements: bool,
) -> tuple[ImportStats, int]:
    stats = ImportStats()
    standard = await get_or_create_standard(session, parsed)
    total_items = 0

    for block in parsed.blocks:
        section = await upsert_section(
            session,
            standard_id=standard.id,
            code=block.disclosure_code,
            title=f"{block.disclosure_code} {block.title}",
            sort_order=block.sort_order,
            parent_section_id=None,
            stats=stats,
        )

        subsection_order = 1
        for subsection in block.subsections:
            if subsection.title == "Requirements" and subsection.row_kind == "requirements":
                continue
            await upsert_section(
                session,
                standard_id=standard.id,
                code=f"{block.disclosure_code}.{slugify(subsection.title)}",
                title=subsection.title,
                sort_order=subsection_order,
                parent_section_id=section.id,
                stats=stats,
            )
            subsection_order += 1

        disclosure = await upsert_disclosure(
            session,
            standard_id=standard.id,
            section_id=section.id,
            block=block,
            stats=stats,
        )

        active_item_codes: set[str] = set()
        for sort_order, (row, datapoint) in enumerate(build_parsed_data_points(block), start=1):
            item = await upsert_esrs_requirement_item(
                session,
                disclosure_id=disclosure.id,
                datapoint=datapoint,
                sort_order=sort_order,
                description=datapoint.label.split(" - ", 1)[-1],
                metadata={
                    "source_ref": row.ref,
                    "clause_ref": row.clause_ref,
                    "row_kind": row.row_kind,
                    "source": row.source,
                    "interpretation": row.interpretation,
                    "evidence": row.evidence,
                    "owner": row.owner,
                    "frequency": row.frequency,
                },
                stats=stats,
            )
            active_item_codes.add(datapoint.item_code)
            total_items += 1
            if with_shared_elements:
                concept_domain = re.sub(r"[^a-z0-9]+", "_", f"{parsed.standard_code}_{block.title}".lower()).strip("_")
                element = await upsert_shared_element(session, item, concept_domain, stats)
                await ensure_mapping(session, item.id, element.id, stats)

        await deactivate_stale_requirement_items(session, disclosure.id, active_item_codes)

    return stats, total_items


def iter_xlsx_paths(path: Path) -> list[Path]:
    if path.is_file():
        if path.suffix.lower() != ".xlsx":
            raise ValueError(f"Expected an .xlsx file, got: {path}")
        return [path]
    if path.is_dir():
        paths = sorted(child for child in path.iterdir() if child.is_file() and child.suffix.lower() == ".xlsx")
        if not paths:
            raise ValueError(f"No .xlsx files found in directory: {path}")
        return paths
    raise ValueError(f"Path does not exist or is not supported: {path}")


async def run_single_import(
    session_factory: async_sessionmaker[AsyncSession],
    path: Path,
    *,
    apply: bool,
    with_shared_elements: bool,
) -> tuple[EsrsParsedWorkbook, ImportStats, int]:
    parsed = parse_workbook(path)
    async with session_factory() as session:
        stats, total_items = await import_workbook(session, parsed, with_shared_elements=with_shared_elements)
        if apply:
            await session.commit()
        else:
            await session.rollback()
    return parsed, stats, total_items


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Import ESRS XLSX workbook(s) into the framework catalog.")
    parser.add_argument(
        "xlsx_path",
        type=Path,
        help="Path to an ESRS .xlsx file or a directory containing ESRS .xlsx files.",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Persist changes. Without this flag, the script runs as a dry-run.",
    )
    parser.add_argument(
        "--no-shared-elements",
        action="store_true",
        help="Skip creating shared elements and mappings.",
    )
    return parser


async def async_main(args: argparse.Namespace) -> int:
    xlsx_paths = iter_xlsx_paths(args.xlsx_path)
    engine = create_async_engine(settings.database_url, echo=False)
    session_factory = async_sessionmaker(engine, expire_on_commit=False)
    aggregate_stats = ImportStats()
    total_disclosures = 0
    total_items = 0

    try:
        print(f"Input: {args.xlsx_path}")
        print(f"Workbooks queued: {len(xlsx_paths)}")
        print(f"Mode: {'apply' if args.apply else 'dry-run'}")
        print()
        for index, xlsx_path in enumerate(xlsx_paths, start=1):
            parsed, stats, workbook_items = await run_single_import(
                session_factory,
                xlsx_path,
                apply=args.apply,
                with_shared_elements=not args.no_shared_elements,
            )
            aggregate_stats.merge(stats)
            total_disclosures += len(parsed.blocks)
            total_items += workbook_items
            print(f"[{index}/{len(xlsx_paths)}] {xlsx_path.name}")
            print(f"  Standard:   {parsed.standard_code} - {parsed.standard_name}")
            print(f"  Blocks:     {len(parsed.blocks)} disclosures")
            print(f"  Datapoints: {workbook_items}")
            print(
                f"  Import summary: sections +{stats.created_sections} / {stats.updated_sections} updated, "
                f"disclosures +{stats.created_disclosures} / {stats.updated_disclosures} updated, "
                f"items +{stats.created_items} / {stats.updated_items} updated"
            )
            print()
    finally:
        await engine.dispose()

    print("Aggregate summary")
    print(f"  Disclosures parsed: {total_disclosures}")
    print(f"  Requirement items parsed: {total_items}")
    print(f"  Sections:         +{aggregate_stats.created_sections} new, {aggregate_stats.updated_sections} updated")
    print(f"  Disclosures:      +{aggregate_stats.created_disclosures} new, {aggregate_stats.updated_disclosures} updated")
    print(f"  RequirementItems: +{aggregate_stats.created_items} new, {aggregate_stats.updated_items} updated")
    print(
        f"  SharedElements:   +{aggregate_stats.created_shared_elements} new, "
        f"{aggregate_stats.updated_shared_elements} updated"
    )
    print(f"  Mappings:         +{aggregate_stats.created_mappings} new")
    return 0


def main() -> int:
    parser = build_arg_parser()
    args = parser.parse_args()
    return asyncio.run(async_main(args))


if __name__ == "__main__":
    raise SystemExit(main())
