"""
Dataset import service.

Handles CSV/XLSX import with auto-detection and preview.
"""

import csv
import io
import re
from typing import Any, Literal

from app.domain.schemas.dataset import ColumnSchema, DatasetImportPreview


class DatasetImportService:
    """Service for importing datasets from files."""

    @staticmethod
    def detect_csv_delimiter(content: bytes, sample_size: int = 5000) -> str:
        """
        Auto-detect CSV delimiter.

        Args:
            content: File content as bytes
            sample_size: Number of bytes to sample

        Returns:
            Detected delimiter: ',' or ';' or '\t'
        """
        sample = content[:sample_size].decode('utf-8', errors='ignore')

        # Count occurrences of common delimiters
        delimiters = {
            ',': sample.count(','),
            ';': sample.count(';'),
            '\t': sample.count('\t'),
        }

        # Return the most common delimiter
        return max(delimiters, key=delimiters.get)

    @staticmethod
    def detect_encoding(content: bytes) -> str:
        """
        Auto-detect file encoding.

        Args:
            content: File content as bytes

        Returns:
            Encoding name: 'utf-8', 'utf-8-sig', 'cp1251', etc.
        """
        # Try UTF-8 with BOM first
        if content.startswith(b'\xef\xbb\xbf'):
            return 'utf-8-sig'

        # Try UTF-8
        try:
            content.decode('utf-8')
            return 'utf-8'
        except UnicodeDecodeError:
            pass

        # Try cp1251 (Cyrillic Windows)
        try:
            content.decode('cp1251')
            return 'cp1251'
        except UnicodeDecodeError:
            pass

        # Fallback to latin1 (always succeeds)
        return 'latin1'

    @staticmethod
    def infer_column_type(values: list[Any]) -> Literal["text", "number", "percent", "date"]:
        """
        Infer column type from sample values.

        Args:
            values: Sample values from column

        Returns:
            Inferred type
        """
        # Filter out nulls for analysis
        non_null = [v for v in values if v is not None and str(v).strip() != '']

        if not non_null:
            return "text"

        # Check if all are numbers
        numeric_count = 0
        percent_count = 0
        date_like_count = 0

        for v in non_null:
            v_str = str(v).strip()

            # Check for percent
            if v_str.endswith('%'):
                percent_count += 1
                v_str = v_str[:-1].strip()

            # Check for date-like patterns before numeric inference
            # - Year: 4 digits, reasonable range
            # - ISO: YYYY-MM or YYYY-MM-DD
            # - Quarter: Q1 2024, Q2-2025, etc.
            try:
                if len(v_str) == 4 and v_str.isdigit():
                    year = int(v_str)
                    if 1800 <= year <= 2200:
                        date_like_count += 1
                elif re.match(r"^\d{4}-\d{2}(-\d{2})?$", v_str):
                    date_like_count += 1
                elif re.match(r"^Q[1-4][\s\-]?\d{4}$", v_str.upper()):
                    date_like_count += 1
            except Exception:
                # Be conservative; ignore parsing errors
                pass

            # Try parse as number
            # Handle thousand separators and decimal point
            v_clean = v_str.replace(' ', '').replace(',', '.')
            try:
                float(v_clean)
                numeric_count += 1
            except ValueError:
                pass

        # If >80% are percentages
        if percent_count > len(non_null) * 0.8:
            return "percent"

        # If >70% are date-like (e.g., years), treat as date BEFORE numeric.
        # This prevents year columns ("2020", "2021", ...) from being inferred as "number".
        if date_like_count > len(non_null) * 0.7:
            return "date"

        # If >80% are numeric
        if numeric_count > len(non_null) * 0.8:
            return "number"

        return "text"

    @staticmethod
    def parse_value(value: str, column_type: str) -> Any:
        """
        Parse value according to column type.

        Args:
            value: Raw string value
            column_type: Type to parse as

        Returns:
            Parsed value (number, string, or None)
        """
        if value is None or str(value).strip() in ('', 'N/A', 'n/a', '—', '-'):
            return None

        v_str = str(value).strip()

        if column_type in ("number", "percent"):
            # Remove percent sign
            if v_str.endswith('%'):
                v_str = v_str[:-1].strip()

            # Handle thousand separators and decimal point
            # Try European format first (comma = decimal, space/dot = thousand)
            v_clean = v_str.replace(' ', '').replace('\u202f', '')  # Remove thin spaces

            # If has both comma and dot, determine which is decimal
            if ',' in v_clean and '.' in v_clean:
                # If comma comes after dot, dot is thousand separator
                if v_clean.rindex(',') > v_clean.rindex('.'):
                    v_clean = v_clean.replace('.', '').replace(',', '.')
                else:
                    v_clean = v_clean.replace(',', '')
            elif ',' in v_clean:
                # Check position: if near end, it's likely decimal
                comma_pos = v_clean.rindex(',')
                if len(v_clean) - comma_pos <= 3:  # Within 2-3 chars of end
                    v_clean = v_clean.replace(',', '.')
                else:
                    v_clean = v_clean.replace(',', '')

            try:
                return float(v_clean)
            except ValueError:
                return None

        # For text and date, return as-is
        return v_str

    async def import_csv_preview(
        self,
        content: bytes,
        filename: str,
    ) -> DatasetImportPreview:
        """
        Preview CSV import with auto-detection.

        Args:
            content: File content
            filename: Original filename

        Returns:
            Import preview with detected schema and sample rows
        """
        encoding = self.detect_encoding(content)
        delimiter = self.detect_csv_delimiter(content)

        text = content.decode(encoding)
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)

        rows = list(reader)

        if not rows:
            return DatasetImportPreview(
                detected_columns=[],
                preview_rows=[],
                total_rows=0,
                warnings=[],
                errors=["File is empty"],
            )

        # First row as headers
        headers = rows[0]
        data_rows = rows[1:]

        if not data_rows:
            return DatasetImportPreview(
                detected_columns=[],
                preview_rows=[],
                total_rows=0,
                warnings=[],
                errors=["No data rows found (only headers)"],
            )

        # Infer column types from first 100 rows
        sample_size = min(100, len(data_rows))
        detected_columns: list[ColumnSchema] = []
        warnings: list[str] = []

        for col_idx, header in enumerate(headers):
            # Get sample values for this column
            sample_values = [
                row[col_idx] if col_idx < len(row) else None
                for row in data_rows[:sample_size]
            ]

            col_type = self.infer_column_type(sample_values)

            detected_columns.append(ColumnSchema(
                key=f"col_{col_idx}",
                type=col_type,
                unit=None,
                nullable=True,
            ))

        # Parse preview rows (first 50)
        preview_rows: list[list] = []
        for row_idx, row in enumerate(data_rows[:50]):
            parsed_row = []
            for col_idx, value in enumerate(row):
                col_type = detected_columns[col_idx].type if col_idx < len(detected_columns) else "text"
                parsed_value = self.parse_value(value, col_type)
                parsed_row.append(parsed_value)
            preview_rows.append(parsed_row)

        # Add warnings
        if len(data_rows) > 10000:
            warnings.append(f"Large dataset: {len(data_rows)} rows. Consider splitting.")

        if len(headers) > 40:
            warnings.append(f"Many columns: {len(headers)}. May affect performance.")

        return DatasetImportPreview(
            detected_columns=detected_columns,
            preview_rows=preview_rows,
            total_rows=len(data_rows),
            warnings=warnings,
            errors=[],
        )

    async def import_xlsx_preview(
        self,
        content: bytes,
        filename: str,
        sheet_name: str | None = None,
        skip_rows: int = 0,
    ) -> DatasetImportPreview:
        """
        Preview XLSX import.

        Args:
            content: File content
            filename: Original filename
            sheet_name: Sheet to import (None = first sheet)
            skip_rows: Number of rows to skip before headers

        Returns:
            Import preview
        """
        try:
            import openpyxl  # type: ignore[import-not-found]
        except Exception:
            return DatasetImportPreview(
                detected_columns=[],
                preview_rows=[],
                total_rows=0,
                warnings=[],
                errors=["XLSX import requires 'openpyxl' (dependency missing)"],
            )

        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

        # Select sheet
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                return DatasetImportPreview(
                    detected_columns=[],
                    preview_rows=[],
                    total_rows=0,
                    warnings=[],
                    errors=[f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}"],
                )
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # Read all rows
        rows = list(ws.iter_rows(values_only=True))

        if not rows or len(rows) <= skip_rows:
            return DatasetImportPreview(
                detected_columns=[],
                preview_rows=[],
                total_rows=0,
                warnings=[],
                errors=["No data rows found after skipping"],
            )

        # Skip specified rows
        rows = rows[skip_rows:]

        # First row as headers
        headers = [str(h) if h is not None else f"Column {i}" for i, h in enumerate(rows[0])]
        data_rows = rows[1:]

        if not data_rows:
            return DatasetImportPreview(
                detected_columns=[],
                preview_rows=[],
                total_rows=0,
                warnings=[],
                errors=["No data rows found (only headers)"],
            )

        # Infer column types
        sample_size = min(100, len(data_rows))
        detected_columns: list[ColumnSchema] = []
        warnings: list[str] = []

        for col_idx, header in enumerate(headers):
            sample_values = [
                row[col_idx] if col_idx < len(row) else None
                for row in data_rows[:sample_size]
            ]

            col_type = self.infer_column_type(sample_values)

            detected_columns.append(ColumnSchema(
                key=f"col_{col_idx}",
                type=col_type,
                unit=None,
                nullable=True,
            ))

        # Parse preview rows
        preview_rows: list[list] = []
        for row in data_rows[:50]:
            parsed_row = []
            for col_idx, value in enumerate(row):
                col_type = detected_columns[col_idx].type if col_idx < len(detected_columns) else "text"

                # XLSX values are already typed by openpyxl
                if value is None:
                    parsed_row.append(None)
                elif isinstance(value, (int, float)):
                    parsed_row.append(value)
                else:
                    parsed_value = self.parse_value(str(value), col_type)
                    parsed_row.append(parsed_value)

            preview_rows.append(parsed_row)

        # Warnings
        if len(data_rows) > 10000:
            warnings.append(f"Large dataset: {len(data_rows)} rows")

        if len(wb.sheetnames) > 1:
            warnings.append(f"Multiple sheets found: {', '.join(wb.sheetnames)}. Currently importing: {ws.title}")

        return DatasetImportPreview(
            detected_columns=detected_columns,
            preview_rows=preview_rows,
            total_rows=len(data_rows),
            warnings=warnings,
            errors=[],
        )

    async def import_csv_full(
        self,
        content: bytes,
        schema_json: dict,
        skip_rows: int = 0,
        max_rows: int | None = None,
    ) -> list[list]:
        """
        Import full CSV with confirmed schema.

        Args:
            content: File content
            schema_json: Final schema (columns)
            skip_rows: Rows to skip after header
            max_rows: Max rows to import

        Returns:
            List of parsed rows
        """
        encoding = self.detect_encoding(content)
        delimiter = self.detect_csv_delimiter(content)

        text = content.decode(encoding)
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)

        rows = list(reader)

        # Skip header + extra rows
        data_rows = rows[1 + skip_rows:]

        # Limit rows
        if max_rows:
            data_rows = data_rows[:max_rows]

        # Parse rows
        columns = schema_json.get("columns", [])
        parsed_rows: list[list] = []

        for row in data_rows:
            parsed_row = []
            for col_idx, col in enumerate(columns):
                value = row[col_idx] if col_idx < len(row) else None
                parsed_value = self.parse_value(value, col["type"])
                parsed_row.append(parsed_value)
            parsed_rows.append(parsed_row)

        return parsed_rows

    async def import_xlsx_full(
        self,
        content: bytes,
        schema_json: dict,
        sheet_name: str | None = None,
        skip_rows: int = 0,
        max_rows: int | None = None,
    ) -> list[list]:
        """
        Import full XLSX with confirmed schema.

        Args:
            content: File content
            schema_json: Final schema
            sheet_name: Sheet name
            skip_rows: Rows to skip after header
            max_rows: Max rows to import

        Returns:
            List of parsed rows
        """
        try:
            import openpyxl  # type: ignore[import-not-found]
        except Exception:
            # Keep behavior explicit; callers can handle and show error in UI.
            raise RuntimeError("XLSX import requires 'openpyxl' (dependency missing)")

        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        rows = list(ws.iter_rows(values_only=True))

        # Skip header + extra rows
        data_rows = rows[1 + skip_rows:]

        # Limit rows
        if max_rows:
            data_rows = data_rows[:max_rows]

        # Parse rows
        columns = schema_json.get("columns", [])
        parsed_rows: list[list] = []

        for row in data_rows:
            parsed_row = []
            for col_idx, col in enumerate(columns):
                value = row[col_idx] if col_idx < len(row) else None

                if value is None:
                    parsed_row.append(None)
                elif isinstance(value, (int, float)):
                    parsed_row.append(value)
                else:
                    parsed_value = self.parse_value(str(value), col["type"])
                    parsed_row.append(parsed_value)

            parsed_rows.append(parsed_row)

        return parsed_rows

