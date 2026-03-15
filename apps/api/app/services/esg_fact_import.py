"""
ESG Fact bulk import service (CSV/XLSX) with preview + confirm.

Design goals:
- No new DB entities (no import jobs table).
- Row-level validation with per-row error reporting.
- Idempotency by content: re-importing the same rows does not create new versions.
"""

from __future__ import annotations

import csv
import io
import json
from contextlib import suppress
from dataclasses import dataclass
from datetime import date
from typing import Any
from uuid import UUID

from fastapi import HTTPException, status
from sqlalchemy import func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.domain.models import Dataset, EsgEntity, EsgFact, EsgFactStatus, EsgLocation, EsgMetric, EsgSegment
from app.domain.models.esg_metric import EsgMetricValueType
from app.domain.schemas.esg_import import (
    EsgFactImportConfirmDTO,
    EsgFactImportPreviewDTO,
    EsgFactImportRowErrorDTO,
    EsgFactImportRowPreviewDTO,
)
from app.services.dataset_import import DatasetImportService
from app.services.esg_logical_key import compute_fact_logical_key_hash, normalize_tags
from app.infra.db_errors import pg_constraint_name, pg_sqlstate

try:
    import openpyxl  # type: ignore[import-not-found]
except Exception:  # pragma: no cover
    openpyxl = None

__all__ = ["EsgFactImportService"]

_PG_UNIQUE_VIOLATION = "23505"
_UQ_FACT_VERSION = "uq_esg_facts_company_logical_version"


def _norm_header(raw: str) -> str:
    h = raw.strip().lower()
    h = h.replace("\ufeff", "")  # BOM
    for ch in (" ", "-", ".", "/"):
        h = h.replace(ch, "_")
    while "__" in h:
        h = h.replace("__", "_")
    return h.strip("_")


def _parse_uuid(raw: str) -> UUID:
    try:
        return UUID(raw)
    except Exception as e:
        raise ValueError("Invalid UUID") from e


def _parse_date(raw: str) -> date:
    try:
        return date.fromisoformat(raw)
    except Exception as e:
        raise ValueError("Invalid date (expected YYYY-MM-DD)") from e


def _parse_bool(raw: str) -> bool:
    v = raw.strip().lower()
    if v in ("true", "1", "yes", "y"):
        return True
    if v in ("false", "0", "no", "n"):
        return False
    raise ValueError("Invalid boolean (expected true/false)")


def _coerce_number(raw: str) -> float:
    parsed = DatasetImportService.parse_value(raw, "number")
    if isinstance(parsed, (int, float)):
        return float(parsed)
    raise ValueError("Invalid number")


def _coerce_int(raw: str) -> int:
    n = _coerce_number(raw)
    if not float(n).is_integer():
        raise ValueError("Invalid integer")
    return int(n)


@dataclass(frozen=True)
class _RawRow:
    row_number: int
    data: dict[str, Any]


@dataclass(frozen=True)
class _ParsedRow:
    row_number: int
    metric: EsgMetric
    metric_code: str | None

    period_type: str
    period_start: date
    period_end: date
    is_ytd: bool

    entity_id: UUID | None
    location_id: UUID | None
    segment_id: UUID | None
    consolidation_approach: str | None
    ghg_scope: str | None
    scope2_method: str | None
    scope3_category: str | None
    tags: list[str] | None

    value_json: Any | None
    dataset_id: UUID | None

    quality_json: dict
    sources_json: dict

    logical_key_hash: str


class EsgFactImportService:
    def __init__(self, session: AsyncSession, *, company_id: UUID, user_id: UUID):
        self.session = session
        self.company_id = company_id
        self.user_id = user_id

        self._metrics_by_code: dict[str, EsgMetric] = {}
        self._metrics_by_id: dict[UUID, EsgMetric] = {}
        self._entities_by_code: dict[str, EsgEntity] = {}
        self._locations_by_code: dict[str, EsgLocation] = {}
        self._segments_by_code: dict[str, EsgSegment] = {}

        self._preloaded = False

    async def _preload_lookups(self) -> None:
        if self._preloaded:
            return

        metrics_stmt = select(EsgMetric).where(EsgMetric.company_id == self.company_id)
        metrics = (await self.session.execute(metrics_stmt)).scalars().all()
        for m in metrics:
            self._metrics_by_id[m.metric_id] = m
            if m.code:
                self._metrics_by_code[m.code.strip().lower()] = m

        entities_stmt = select(EsgEntity).where(EsgEntity.company_id == self.company_id)
        for e in (await self.session.execute(entities_stmt)).scalars().all():
            if e.code:
                self._entities_by_code[e.code.strip().lower()] = e

        locations_stmt = select(EsgLocation).where(EsgLocation.company_id == self.company_id)
        for l in (await self.session.execute(locations_stmt)).scalars().all():
            if l.code:
                self._locations_by_code[l.code.strip().lower()] = l

        segments_stmt = select(EsgSegment).where(EsgSegment.company_id == self.company_id)
        for s in (await self.session.execute(segments_stmt)).scalars().all():
            if s.code:
                self._segments_by_code[s.code.strip().lower()] = s

        self._preloaded = True

    async def _get_metric(self, *, metric_id: UUID | None, metric_code: str | None) -> EsgMetric:
        await self._preload_lookups()

        if metric_id is not None:
            m = self._metrics_by_id.get(metric_id)
            if m:
                return m
            # Fallback (should not happen with preload, but keep explicit).
            m = await self.session.get(EsgMetric, metric_id)
            if not m or m.company_id != self.company_id:
                raise ValueError("Metric not found")
            self._metrics_by_id[m.metric_id] = m
            if m.code:
                self._metrics_by_code[m.code.strip().lower()] = m
            return m

        if metric_code:
            m = self._metrics_by_code.get(metric_code.strip().lower())
            if m:
                return m
        raise ValueError("metric_code or metric_id is required")

    async def _resolve_dimension_id(
        self,
        *,
        kind: str,
        raw_id: str | None,
        raw_code: str | None,
    ) -> UUID | None:
        if raw_id and raw_id.strip():
            return _parse_uuid(raw_id.strip())

        if not raw_code or not raw_code.strip():
            return None

        await self._preload_lookups()
        code = raw_code.strip().lower()

        if kind == "entity":
            item = self._entities_by_code.get(code)
            if item:
                return item.entity_id
        elif kind == "location":
            item = self._locations_by_code.get(code)
            if item:
                return item.location_id
        elif kind == "segment":
            item = self._segments_by_code.get(code)
            if item:
                return item.segment_id
        else:
            raise ValueError("Unsupported dimension kind")

        raise ValueError(f"{kind} not found for code '{raw_code.strip()}'")

    async def _ensure_dataset_belongs(self, dataset_id: UUID) -> None:
        ds = await self.session.get(Dataset, dataset_id)
        if not ds or ds.is_deleted or ds.company_id != self.company_id:
            raise ValueError("Dataset not found")

    def _validate_value_for_metric(self, metric: EsgMetric, *, raw_value: Any | None, raw_dataset_id: UUID | None) -> Any | None:
        if metric.value_type == EsgMetricValueType.DATASET:
            if raw_dataset_id is None:
                raise ValueError("dataset_id is required for dataset metrics")
            if raw_value is not None:
                raise ValueError("value must be empty for dataset metrics")
            return None

        if raw_dataset_id is not None:
            raise ValueError("dataset_id must be empty for scalar metrics")

        if raw_value is None:
            raise ValueError("value is required for scalar metrics")

        # XLSX values may already be typed (int/float/bool); CSV is string.
        if metric.value_type == EsgMetricValueType.BOOLEAN:
            if isinstance(raw_value, bool):
                return raw_value
            if isinstance(raw_value, str):
                return _parse_bool(raw_value)
            raise ValueError("Expected boolean value")

        if metric.value_type == EsgMetricValueType.INTEGER:
            if isinstance(raw_value, bool):
                raise ValueError("Invalid integer value")
            if isinstance(raw_value, int):
                return raw_value
            if isinstance(raw_value, float) and raw_value.is_integer():
                return int(raw_value)
            if isinstance(raw_value, str):
                return _coerce_int(raw_value)
            raise ValueError("Expected integer value")

        if metric.value_type == EsgMetricValueType.NUMBER:
            if isinstance(raw_value, bool):
                raise ValueError("Invalid numeric value")
            if isinstance(raw_value, (int, float)):
                return float(raw_value)
            if isinstance(raw_value, str):
                return _coerce_number(raw_value)
            raise ValueError("Expected numeric value")

        if metric.value_type == EsgMetricValueType.STRING:
            if isinstance(raw_value, str):
                return raw_value
            return str(raw_value)

        raise ValueError("Unsupported metric value_type")

    def _parse_tags(self, raw: str | None) -> list[str] | None:
        if not raw or not raw.strip():
            return None
        parts = [p.strip() for p in raw.split(",")]
        tags = normalize_tags([p for p in parts if p])
        return tags or None

    def _parse_json_dict(self, raw: str | None, *, field: str) -> dict:
        if raw is None:
            return {}
        s = raw.strip()
        if not s:
            return {}
        try:
            parsed = json.loads(s)
        except Exception as e:
            raise ValueError(f"Invalid {field} (expected JSON object)") from e
        if not isinstance(parsed, dict):
            raise ValueError(f"Invalid {field} (expected JSON object)")
        return parsed

    async def _parse_row(self, row_number: int, row: dict[str, Any]) -> _ParsedRow:
        metric_code = (row.get("metric_code") or None)
        if isinstance(metric_code, str):
            metric_code = metric_code.strip() or None
        else:
            metric_code = None

        metric_id: UUID | None = None
        raw_metric_id = row.get("metric_id")
        if isinstance(raw_metric_id, str) and raw_metric_id.strip():
            metric_id = _parse_uuid(raw_metric_id.strip())

        metric = await self._get_metric(metric_id=metric_id, metric_code=metric_code)

        raw_period_type = row.get("period_type")
        period_type = str(raw_period_type).strip().lower() if raw_period_type is not None else "year"
        if period_type not in ("day", "month", "quarter", "year", "custom"):
            raise ValueError("Invalid period_type")

        raw_period_start = row.get("period_start")
        raw_period_end = row.get("period_end")
        if not isinstance(raw_period_start, str) or not raw_period_start.strip():
            raise ValueError("period_start is required")
        if not isinstance(raw_period_end, str) or not raw_period_end.strip():
            raise ValueError("period_end is required")
        period_start = _parse_date(raw_period_start.strip())
        period_end = _parse_date(raw_period_end.strip())
        if period_start > period_end:
            raise ValueError("period_start must be <= period_end")

        raw_is_ytd = row.get("is_ytd")
        is_ytd = False
        if isinstance(raw_is_ytd, str) and raw_is_ytd.strip():
            is_ytd = _parse_bool(raw_is_ytd)
        elif isinstance(raw_is_ytd, bool):
            is_ytd = raw_is_ytd

        entity_id = await self._resolve_dimension_id(
            kind="entity",
            raw_id=row.get("entity_id") if isinstance(row.get("entity_id"), str) else None,
            raw_code=row.get("entity_code") if isinstance(row.get("entity_code"), str) else None,
        )
        location_id = await self._resolve_dimension_id(
            kind="location",
            raw_id=row.get("location_id") if isinstance(row.get("location_id"), str) else None,
            raw_code=row.get("location_code") if isinstance(row.get("location_code"), str) else None,
        )
        segment_id = await self._resolve_dimension_id(
            kind="segment",
            raw_id=row.get("segment_id") if isinstance(row.get("segment_id"), str) else None,
            raw_code=row.get("segment_code") if isinstance(row.get("segment_code"), str) else None,
        )

        consolidation_approach = str(row.get("consolidation_approach")).strip() if row.get("consolidation_approach") is not None else ""
        consolidation_approach = consolidation_approach or None
        ghg_scope = str(row.get("ghg_scope")).strip() if row.get("ghg_scope") is not None else ""
        ghg_scope = ghg_scope or None
        scope2_method = str(row.get("scope2_method")).strip() if row.get("scope2_method") is not None else ""
        scope2_method = scope2_method or None
        scope3_category = str(row.get("scope3_category")).strip() if row.get("scope3_category") is not None else ""
        scope3_category = scope3_category or None

        tags = self._parse_tags(row.get("tags") if isinstance(row.get("tags"), str) else None)

        raw_dataset_id: UUID | None = None
        raw_ds = row.get("dataset_id")
        if isinstance(raw_ds, str) and raw_ds.strip():
            raw_dataset_id = _parse_uuid(raw_ds.strip())

        raw_value = row.get("value")
        # Empty strings should be treated as missing
        if isinstance(raw_value, str) and not raw_value.strip():
            raw_value = None

        value_json = self._validate_value_for_metric(metric, raw_value=raw_value, raw_dataset_id=raw_dataset_id)

        if raw_dataset_id is not None:
            await self._ensure_dataset_belongs(raw_dataset_id)

        quality_json = self._parse_json_dict(row.get("quality_json") if isinstance(row.get("quality_json"), str) else None, field="quality_json")
        sources_json = self._parse_json_dict(row.get("sources_json") if isinstance(row.get("sources_json"), str) else None, field="sources_json")

        logical_key_hash = compute_fact_logical_key_hash(
            metric_id=metric.metric_id,
            period_start=period_start,
            period_end=period_end,
            period_type=period_type,
            is_ytd=is_ytd,
            entity_id=entity_id,
            location_id=location_id,
            segment_id=segment_id,
            consolidation_approach=consolidation_approach,
            ghg_scope=ghg_scope,
            scope2_method=scope2_method,
            scope3_category=scope3_category,
            tags=tags,
        )

        return _ParsedRow(
            row_number=row_number,
            metric=metric,
            metric_code=metric_code,
            period_type=period_type,
            period_start=period_start,
            period_end=period_end,
            is_ytd=is_ytd,
            entity_id=entity_id,
            location_id=location_id,
            segment_id=segment_id,
            consolidation_approach=consolidation_approach,
            ghg_scope=ghg_scope,
            scope2_method=scope2_method,
            scope3_category=scope3_category,
            tags=tags,
            value_json=value_json,
            dataset_id=raw_dataset_id,
            quality_json=quality_json,
            sources_json=sources_json,
            logical_key_hash=logical_key_hash,
        )

    async def _fetch_latest_facts(self, logical_key_hashes: list[str]) -> dict[str, EsgFact]:
        if not logical_key_hashes:
            return {}

        ranked = (
            select(
                EsgFact.fact_id.label("fact_id"),
                EsgFact.logical_key_hash.label("logical_key_hash"),
                func.row_number()
                .over(
                    partition_by=EsgFact.logical_key_hash,
                    order_by=[EsgFact.version_number.desc()],
                )
                .label("rn"),
            )
            .where(
                EsgFact.company_id == self.company_id,
                EsgFact.logical_key_hash.in_(logical_key_hashes),
            )
            .cte("ranked_import_latest")
        )

        stmt = (
            select(EsgFact)
            .join(ranked, ranked.c.fact_id == EsgFact.fact_id)
            .where(ranked.c.rn == 1)
        )
        facts = (await self.session.execute(stmt)).scalars().all()
        return {f.logical_key_hash: f for f in facts}

    def _facts_equal_for_idempotency(self, latest: EsgFact, incoming: _ParsedRow) -> bool:
        if latest.dataset_id is not None or incoming.dataset_id is not None:
            if latest.dataset_id != incoming.dataset_id:
                return False
        else:
            if latest.value_json != incoming.value_json:
                return False

        return latest.quality_json == incoming.quality_json and latest.sources_json == incoming.sources_json

    async def preview_rows(
        self,
        rows: list[_ParsedRow],
        *,
        total_rows: int,
        file_duplicates: set[str],
    ) -> EsgFactImportPreviewDTO:
        latest_by_hash = await self._fetch_latest_facts([r.logical_key_hash for r in rows])

        previews: list[EsgFactImportRowPreviewDTO] = []
        errors: list[EsgFactImportRowErrorDTO] = []

        create_rows = 0
        skip_rows = 0
        error_rows = 0

        for r in rows:
            if r.logical_key_hash in file_duplicates:
                error_rows += 1
                msg = "Duplicate logical key within file"
                errors.append(
                    EsgFactImportRowErrorDTO(
                        row_number=r.row_number,
                        message=msg,
                        metric_code=r.metric_code,
                        logical_key_hash=r.logical_key_hash,
                    )
                )
                previews.append(
                    EsgFactImportRowPreviewDTO(
                        row_number=r.row_number,
                        action="error",
                        message=msg,
                        metric_code=r.metric_code,
                        logical_key_hash=r.logical_key_hash,
                    )
                )
                continue

            latest = latest_by_hash.get(r.logical_key_hash)
            if latest and latest.status == EsgFactStatus.IN_REVIEW:
                error_rows += 1
                msg = "Latest fact is in_review; request changes before importing"
                errors.append(
                    EsgFactImportRowErrorDTO(
                        row_number=r.row_number,
                        message=msg,
                        metric_code=r.metric_code,
                        logical_key_hash=r.logical_key_hash,
                    )
                )
                previews.append(
                    EsgFactImportRowPreviewDTO(
                        row_number=r.row_number,
                        action="error",
                        message=msg,
                        metric_code=r.metric_code,
                        logical_key_hash=r.logical_key_hash,
                    )
                )
                continue

            if latest and self._facts_equal_for_idempotency(latest, r):
                skip_rows += 1
                previews.append(
                    EsgFactImportRowPreviewDTO(
                        row_number=r.row_number,
                        action="skip",
                        message="No changes (idempotent)",
                        metric_code=r.metric_code,
                        logical_key_hash=r.logical_key_hash,
                    )
                )
                continue

            create_rows += 1
            previews.append(
                EsgFactImportRowPreviewDTO(
                    row_number=r.row_number,
                    action="create",
                    metric_code=r.metric_code,
                    logical_key_hash=r.logical_key_hash,
                )
            )

        return EsgFactImportPreviewDTO(
            total_rows=total_rows,
            create_rows=create_rows,
            skip_rows=skip_rows,
            error_rows=error_rows,
            rows=previews,
            errors=errors,
        )

    async def confirm_rows(
        self,
        rows: list[_ParsedRow],
        *,
        total_rows: int,
        file_duplicates: set[str],
    ) -> EsgFactImportConfirmDTO:
        # Re-check latest at confirm time to preserve idempotency.
        latest_by_hash = await self._fetch_latest_facts([r.logical_key_hash for r in rows])

        created = 0
        skipped = 0
        errors: list[EsgFactImportRowErrorDTO] = []

        for r in rows:
            if r.logical_key_hash in file_duplicates:
                errors.append(
                    EsgFactImportRowErrorDTO(
                        row_number=r.row_number,
                        message="Duplicate logical key within file",
                        metric_code=r.metric_code,
                        logical_key_hash=r.logical_key_hash,
                    )
                )
                continue

            latest = latest_by_hash.get(r.logical_key_hash)
            if latest and latest.status == EsgFactStatus.IN_REVIEW:
                errors.append(
                    EsgFactImportRowErrorDTO(
                        row_number=r.row_number,
                        message="Latest fact is in_review; request changes before importing",
                        metric_code=r.metric_code,
                        logical_key_hash=r.logical_key_hash,
                    )
                )
                continue

            if latest and self._facts_equal_for_idempotency(latest, r):
                skipped += 1
                continue

            # Create a new draft version.
            max_retries = 2
            created_this_row = False

            for attempt in range(max_retries + 1):
                version_number = (latest.version_number + 1) if latest else 1
                supersedes_fact_id = latest.fact_id if latest else None

                fact = EsgFact(
                    company_id=self.company_id,
                    metric_id=r.metric.metric_id,
                    status=EsgFactStatus.DRAFT,
                    version_number=version_number,
                    supersedes_fact_id=supersedes_fact_id,
                    logical_key_hash=r.logical_key_hash,
                    period_type=r.period_type,
                    period_start=r.period_start,
                    period_end=r.period_end,
                    is_ytd=r.is_ytd,
                    entity_id=r.entity_id,
                    location_id=r.location_id,
                    segment_id=r.segment_id,
                    consolidation_approach=r.consolidation_approach,
                    ghg_scope=r.ghg_scope,
                    scope2_method=r.scope2_method,
                    scope3_category=r.scope3_category,
                    tags=r.tags or None,
                    value_json=r.value_json if r.dataset_id is None else None,
                    dataset_id=r.dataset_id,
                    dataset_revision_id=None,
                    quality_json=r.quality_json,
                    sources_json=r.sources_json,
                    created_by=self.user_id,
                    updated_by=self.user_id,
                )

                try:
                    async with self.session.begin_nested():
                        self.session.add(fact)
                        await self.session.flush()
                    created += 1
                    created_this_row = True
                    # Keep cache consistent for any later logic in this request.
                    latest_by_hash[r.logical_key_hash] = fact
                    break
                except IntegrityError as e:
                    with suppress(Exception):
                        self.session.expunge(fact)

                    if pg_sqlstate(e) == _PG_UNIQUE_VIOLATION and pg_constraint_name(e) == _UQ_FACT_VERSION:
                        # Another writer created the same version concurrently. Re-fetch and retry.
                        if attempt < max_retries:
                            refreshed = await self._fetch_latest_facts([r.logical_key_hash])
                            latest = refreshed.get(r.logical_key_hash)
                            if latest is not None:
                                latest_by_hash[r.logical_key_hash] = latest
                            continue

                        errors.append(
                            EsgFactImportRowErrorDTO(
                                row_number=r.row_number,
                                message="Version conflict while importing this row. Please retry.",
                                metric_code=r.metric_code,
                                logical_key_hash=r.logical_key_hash,
                            )
                        )
                        break
                    raise

        await self.session.commit()

        return EsgFactImportConfirmDTO(
            total_rows=total_rows,
            created=created,
            skipped=skipped,
            error_rows=len(errors),
            errors=errors,
        )

    # ---------------------------------------------------------------------
    # CSV/XLSX entry points
    # ---------------------------------------------------------------------

    async def preview_csv(self, *, content: bytes, skip_rows: int = 0) -> EsgFactImportPreviewDTO:
        raw_rows = self._read_csv(content=content, skip_rows=skip_rows)
        parsed, parse_errors, dups = await self._parse_and_validate_rows(raw_rows)
        preview = await self.preview_rows(parsed, total_rows=len(raw_rows), file_duplicates=dups)
        # Include parse-time errors.
        preview.errors.extend(parse_errors)
        preview.error_rows += len(parse_errors)
        return preview

    async def confirm_csv(self, *, content: bytes, skip_rows: int = 0) -> EsgFactImportConfirmDTO:
        raw_rows = self._read_csv(content=content, skip_rows=skip_rows)
        parsed, parse_errors, dups = await self._parse_and_validate_rows(raw_rows)
        result = await self.confirm_rows(parsed, total_rows=len(raw_rows), file_duplicates=dups)
        result.errors.extend(parse_errors)
        result.error_rows += len(parse_errors)
        return result

    async def preview_xlsx(self, *, content: bytes, sheet_name: str | None = None, skip_rows: int = 0) -> EsgFactImportPreviewDTO:
        raw_rows = self._read_xlsx(content=content, sheet_name=sheet_name, skip_rows=skip_rows)
        parsed, parse_errors, dups = await self._parse_and_validate_rows(raw_rows)
        preview = await self.preview_rows(parsed, total_rows=len(raw_rows), file_duplicates=dups)
        preview.errors.extend(parse_errors)
        preview.error_rows += len(parse_errors)
        return preview

    async def confirm_xlsx(self, *, content: bytes, sheet_name: str | None = None, skip_rows: int = 0) -> EsgFactImportConfirmDTO:
        raw_rows = self._read_xlsx(content=content, sheet_name=sheet_name, skip_rows=skip_rows)
        parsed, parse_errors, dups = await self._parse_and_validate_rows(raw_rows)
        result = await self.confirm_rows(parsed, total_rows=len(raw_rows), file_duplicates=dups)
        result.errors.extend(parse_errors)
        result.error_rows += len(parse_errors)
        return result

    def _read_csv(self, *, content: bytes, skip_rows: int) -> list[_RawRow]:
        encoding = DatasetImportService.detect_encoding(content)
        delimiter = DatasetImportService.detect_csv_delimiter(content)
        text = content.decode(encoding)
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        all_rows = list(reader)
        if not all_rows:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File is empty")
        if skip_rows < 0:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="skip_rows must be >= 0")
        if len(all_rows) <= skip_rows:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No header row found after skipping")

        header = all_rows[skip_rows]
        keys = [_norm_header(str(h) if h is not None else "") for h in header]
        data_rows = all_rows[skip_rows + 1 :]

        out: list[_RawRow] = []
        for idx, row in enumerate(data_rows):
            item: dict[str, Any] = {}
            for i, k in enumerate(keys):
                if not k:
                    continue
                item[k] = row[i] if i < len(row) else None
            out.append(_RawRow(row_number=skip_rows + 2 + idx, data=item))
        return out

    def _read_xlsx(self, *, content: bytes, sheet_name: str | None, skip_rows: int) -> list[_RawRow]:
        if openpyxl is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="XLSX import requires 'openpyxl' (dependency missing)",
            )

        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}",
                )
            ws = wb[sheet_name]
        else:
            ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File is empty")
        if skip_rows < 0:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="skip_rows must be >= 0")
        if len(rows) <= skip_rows:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No header row found after skipping")

        header = rows[skip_rows]
        keys = [_norm_header(str(h) if h is not None else "") for h in header]
        data_rows = rows[skip_rows + 1 :]

        out: list[_RawRow] = []
        for idx, row in enumerate(data_rows):
            item: dict[str, Any] = {}
            for i, k in enumerate(keys):
                if not k:
                    continue
                v = row[i] if i < len(row) else None
                item[k] = v
            out.append(_RawRow(row_number=skip_rows + 2 + idx, data=item))
        return out

    async def _parse_and_validate_rows(
        self, raw_rows: list[_RawRow]
    ) -> tuple[list[_ParsedRow], list[EsgFactImportRowErrorDTO], set[str]]:
        parsed: list[_ParsedRow] = []
        errors: list[EsgFactImportRowErrorDTO] = []

        seen_hashes: dict[str, int] = {}
        duplicates: set[str] = set()

        for raw in raw_rows:
            row_number = raw.row_number
            row = raw.data
            try:
                pr = await self._parse_row(row_number, row)
            except Exception as e:
                errors.append(
                    EsgFactImportRowErrorDTO(
                        row_number=row_number,
                        message=str(e),
                        metric_code=row.get("metric_code") if isinstance(row.get("metric_code"), str) else None,
                    )
                )
                continue

            first_seen = seen_hashes.get(pr.logical_key_hash)
            if first_seen is None:
                seen_hashes[pr.logical_key_hash] = pr.row_number
            else:
                duplicates.add(pr.logical_key_hash)

            parsed.append(pr)

        return parsed, errors, duplicates
