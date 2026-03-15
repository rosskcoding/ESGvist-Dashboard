"""
Dataset export service.

Handles CSV/XLSX/JSON export.
"""

import csv
import io
import json
from typing import Literal

from app.domain.models import Dataset, DatasetRevision


class DatasetExportService:
    """Service for exporting datasets to files."""

    @staticmethod
    def export_to_csv(
        dataset: Dataset | DatasetRevision,
        include_metadata: bool = False,
    ) -> bytes:
        """
        Export dataset to CSV format.

        Args:
            dataset: Dataset or DatasetRevision instance
            include_metadata: Include meta_json as comments (not supported in CSV)

        Returns:
            CSV file content as bytes
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Write headers
        columns = dataset.schema_json.get("columns", [])
        if not columns:
            return b""

        # Extract column keys for header row
        # But use human-readable names if available in meta
        headers = []
        for col in columns:
            # Try to get a nice label from meta or use key
            headers.append(col.get("key", ""))

        writer.writerow(headers)

        # Write rows
        rows = dataset.rows_json or []
        for row in rows:
            writer.writerow(row)

        # Convert to bytes
        return output.getvalue().encode('utf-8')

    @staticmethod
    def export_to_xlsx(
        dataset: Dataset | DatasetRevision,
        include_metadata: bool = True,
    ) -> bytes:
        """
        Export dataset to XLSX format.

        Args:
            dataset: Dataset or DatasetRevision instance
            include_metadata: Include metadata in separate sheet

        Returns:
            XLSX file content as bytes
        """
        try:
            import openpyxl  # type: ignore[import-not-found]
            from openpyxl.styles import Font  # type: ignore[import-not-found]
        except Exception as e:
            raise RuntimeError("XLSX export requires 'openpyxl' (dependency missing)") from e

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"

        # Write headers
        columns = dataset.schema_json.get("columns", [])
        if not columns:
            # Empty workbook
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()

        headers = [col.get("key", "") for col in columns]
        ws.append(headers)

        # Style header row
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Write rows
        rows = dataset.rows_json or []
        for row in rows:
            ws.append(row)

        # Auto-size columns
        for col_idx, col in enumerate(columns, start=1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 15

        # Add metadata sheet if requested
        if include_metadata:
            meta_ws = wb.create_sheet("Metadata")
            meta_ws.append(["Key", "Value"])

            # Dataset info
            if isinstance(dataset, Dataset):
                meta_ws.append(["Dataset ID", str(dataset.dataset_id)])
                meta_ws.append(["Name", dataset.name])
                meta_ws.append(["Description", dataset.description or ""])
                meta_ws.append(["Current Revision", dataset.current_revision])
                meta_ws.append(["Created At", dataset.created_at_utc.isoformat()])
                meta_ws.append(["Updated At", dataset.updated_at_utc.isoformat()])
            else:
                meta_ws.append(["Revision ID", str(dataset.revision_id)])
                meta_ws.append(["Revision Number", dataset.revision_number])
                meta_ws.append(["Created At", dataset.created_at_utc.isoformat()])
                meta_ws.append(["Reason", dataset.reason or ""])

            # Custom metadata from meta_json
            meta_json = dataset.meta_json or {}
            for key, value in meta_json.items():
                meta_ws.append([key, str(value)])

            # Style metadata header
            for cell in meta_ws[1]:
                cell.font = Font(bold=True)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def export_to_json(
        dataset: Dataset | DatasetRevision,
        include_metadata: bool = True,
        format_type: Literal["row_based", "column_based"] = "row_based",
    ) -> bytes:
        """
        Export dataset to JSON format.

        Args:
            dataset: Dataset or DatasetRevision instance
            include_metadata: Include schema and metadata
            format_type: row_based [{col: val, ...}] or column_based {columns, rows}

        Returns:
            JSON file content as bytes
        """
        columns = dataset.schema_json.get("columns", [])
        rows = dataset.rows_json or []

        if format_type == "row_based":
            # [{col1: val1, col2: val2}, ...]
            data = []
            for row in rows:
                row_dict = {}
                for col_idx, col in enumerate(columns):
                    key = col.get("key", f"col_{col_idx}")
                    value = row[col_idx] if col_idx < len(row) else None
                    row_dict[key] = value
                data.append(row_dict)

            if include_metadata:
                result = {
                    "metadata": _extract_metadata(dataset),
                    "schema": dataset.schema_json,
                    "data": data,
                }
            else:
                result = data

        else:  # column_based
            # {columns: [...], rows: [[...]]}
            if include_metadata:
                result = {
                    "metadata": _extract_metadata(dataset),
                    "schema": dataset.schema_json,
                    "rows": rows,
                }
            else:
                result = {
                    "columns": [col.get("key", "") for col in columns],
                    "rows": rows,
                }

        return json.dumps(result, indent=2, ensure_ascii=False).encode('utf-8')


def _extract_metadata(dataset: Dataset | DatasetRevision) -> dict:
    """Extract metadata from Dataset or DatasetRevision."""
    if isinstance(dataset, Dataset):
        return {
            "dataset_id": str(dataset.dataset_id),
            "name": dataset.name,
            "description": dataset.description,
            "current_revision": dataset.current_revision,
            "created_at_utc": dataset.created_at_utc.isoformat(),
            "updated_at_utc": dataset.updated_at_utc.isoformat(),
            "custom": dataset.meta_json,
        }
    else:
        return {
            "revision_id": str(dataset.revision_id),
            "dataset_id": str(dataset.dataset_id),
            "revision_number": dataset.revision_number,
            "created_at_utc": dataset.created_at_utc.isoformat(),
            "reason": dataset.reason,
            "custom": dataset.meta_json,
        }

